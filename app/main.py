from __future__ import annotations

import io
import os
import shutil
import sqlite3
import threading
import time
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Optional, Tuple, List, Dict

from collections import deque
import requests
from fastapi import FastAPI, Request, Form, UploadFile, File
from fastapi.responses import RedirectResponse, StreamingResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from watchdog.events import FileSystemEventHandler
from watchdog.observers import Observer
from openpyxl import Workbook, load_workbook

APP_NAME = "PitiPiti - Mükellef Bildirimi"
DB_PATH = Path("storage/app.db")
WA_SENDER_BASE = "http://127.0.0.1:3210"
TR_TZ = timezone(timedelta(hours=3))


# ===============================
# SEND concurrency guard (single active send)
# ===============================
SEND_MUTEX = threading.Lock()

def _queue_company_send(company_id: int, when_utc_iso: str, total_files: int, note: str = "Kuyruğa alındı.") -> None:
    """Gönderim meşgulse işi kuyruğa al: next_send_at + total file sayısı."""
    try:
        ensure_company_job_cols()
        con = db()
        cur = con.cursor()
        cur.execute(
            """
            UPDATE companies
            SET next_send_at=?,
                next_send_total_files=?,
                last_send_status=?,
                last_error=?
            WHERE id=?
            """,
            (when_utc_iso, int(total_files or 0), "QUEUED", note, int(company_id)),
        )
        con.commit()
        con.close()
    except Exception:
        try:
            con.close()
        except Exception:
            pass






def ensure_company_job_cols():
    try:
        con = db()
        cur = con.cursor()
        cur.execute("PRAGMA table_info(companies)")
        cols = {r[1] for r in cur.fetchall()}

        alters = []
        if "next_send_total_files" not in cols:
            alters.append("ALTER TABLE companies ADD COLUMN next_send_total_files INTEGER")
        if "last_send_status_text" not in cols:
            alters.append("ALTER TABLE companies ADD COLUMN last_send_status_text TEXT")
        if "last_send_extra_count" not in cols:
            alters.append("ALTER TABLE companies ADD COLUMN last_send_extra_count INTEGER")

        for sql in alters:
            cur.execute(sql)

        con.commit()
        con.close()
    except Exception:
        try:
            con.close()
        except Exception:
            pass




# ===============================
# SIDEBAR İŞ TAKİBİ (Planned + Done, 5sn TTL)
# ===============================
RECENT_JOBS = deque(maxlen=80)  # {"ts":float,"company_id":int,"name":str,"x":int,"y":int,"type":"done"}
RECENT_TTL_SEC = 5

def add_recent_job(company_id: int, name: str, x: int, y: int) -> None:
    """Tamamlanan işi sidebar için ekler (5 sn görünür)."""
    try:
        RECENT_JOBS.appendleft({
            "ts": time.time(),
            "company_id": int(company_id),
            "name": (name or "-"),
            "x": int(x or 0),
            "y": int(y or 0),
            "type": "done",
        })
    except Exception:
        pass

def _fmt_tr_ddmmyyyy_hhmm(dt: Optional[datetime]) -> str:
    if not dt:
        return ""
    try:
        return dt.astimezone(TR_TZ).strftime("%d.%m.%Y %H:%M")
    except Exception:
        try:
            return dt.strftime("%d.%m.%Y %H:%M")
        except Exception:
            return ""

def get_planned_jobs(limit: int = 10) -> list[dict]:
    """
    Planlanan işleri DB'den çeker:
      - SADECE gelecekteki next_send_at (UTC) döner
      - geçmiş planları otomatik temizler
      - y (toplam evrak): companies.next_send_total_files varsa onu kullanır, yoksa READY dosyaları sayar
    """
    ensure_company_job_cols()

    # 0) geçmiş planları temizle (SENDING olanlara dokunma)
    try:
        con = db()
        cur = con.cursor()
        cur.execute(
            """
            UPDATE companies
            SET next_send_at='',
                next_send_total_files=0
            WHERE COALESCE(next_send_at,'') <> ''
              AND datetime(REPLACE(substr(next_send_at,1,19),'T',' ')) < datetime('now')
              AND COALESCE(last_send_status,'') <> 'SENDING'
            """
        )
        con.commit()
        con.close()
    except Exception:
        try:
            con.close()
        except Exception:
            pass

    # 1) sadece gelecekteki planları çek
    con = db()
    cur = con.cursor()
    cur.execute(
        """
        SELECT id, name, COALESCE(next_send_at,''), COALESCE(next_send_total_files, NULL)
        FROM companies
        WHERE COALESCE(is_active,1)=1
          AND TRIM(COALESCE(next_send_at,'')) <> ''
          AND datetime(REPLACE(substr(next_send_at,1,19),'T',' ')) >= datetime('now')
        ORDER BY
          datetime(REPLACE(substr(next_send_at,1,19),'T',' ')) ASC,
          id DESC
        LIMIT ?
        """,
        (int(limit),),
    )
    rows = cur.fetchall()
    con.close()

    items: list[dict] = []
    for cid, name, next_send_at, total_files in rows:
        # y (toplam evrak)
        y = 0
        if total_files is not None:
            try:
                y = int(total_files or 0)
            except Exception:
                y = 0
        else:
            try:
                rr = documents_ready_rows(int(cid))
                y = sum(1 for _, p in rr if p and Path(p).exists())
            except Exception:
                y = 0

        items.append(
            {
                "company_id": int(cid),
                "name": name or "-",
                "x": 0,
                "y": int(y),
                "next_send_at": (next_send_at or "").strip(),
                "type": "planned",
            }
        )

    return items

def utcnow() -> datetime:
    return datetime.now(timezone.utc)


def local_now() -> datetime:
    return datetime.now(TR_TZ)


def iso(dt: datetime) -> str:
    return dt.astimezone(timezone.utc).isoformat()


def parse_iso(s: str) -> Optional[datetime]:
    """ISO tarih parse eder. Naive gelirse UTC kabul edip tz ekler."""
    try:
        dt = datetime.fromisoformat((s or "").strip())
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=UTC)
        return dt
    except Exception:
        return None


def parse_dt_local_to_iso(dt_local: str) -> str:
    """
    dt_local: HTML datetime-local -> 'YYYY-MM-DDTHH:MM' (bazen saniyesiz)
    return: ISO string (naive) 'YYYY-MM-DDTHH:MM:SS'
    """
    s = (dt_local or "").strip()
    if not s:
        raise ValueError("Tarih/Saat boş.")
    # bazen 'YYYY-MM-DD HH:MM' gelebilir
    s = s.replace(" ", "T")
    # saniye yoksa ekle
    if len(s) == 16:
        s = s + ":00"
    # fromisoformat bunu kabul eder (timezone yok)
    datetime.fromisoformat(s)
    return s

def db() -> sqlite3.Connection:
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    con = sqlite3.connect(DB_PATH, timeout=10, check_same_thread=False)
    con.execute("PRAGMA busy_timeout=10000;")
    con.execute("PRAGMA journal_mode=WAL;")
    con.execute("PRAGMA synchronous=NORMAL;")
    return con


def ensure_table_columns(table: str, columns: dict) -> None:
    """SQLite tablo kolonlarını güvenli şekilde tamamlar."""
    con = db()
    cur = con.cursor()
    try:
        cur.execute(f"PRAGMA table_info({table})")
        existing = {r[1] for r in cur.fetchall()}
    except Exception:
        existing = set()
    for col, ddl in columns.items():
        if col not in existing:
            try:
                cur.execute(ddl)
            except Exception:
                pass
    con.commit()
    con.close()



def _table_cols(cur: sqlite3.Cursor, table: str) -> set[str]:
    cur.execute(f"PRAGMA table_info({table})")
    return {r[1] for r in cur.fetchall()}


def ensure_db() -> None:
    con = db()
    cur = con.cursor()

    cur.execute(
        """
    CREATE TABLE IF NOT EXISTS settings (
        key TEXT PRIMARY KEY,
        value TEXT NOT NULL
    )"""
    )

    cur.execute(
        """
    CREATE TABLE IF NOT EXISTS license_state (
        id INTEGER PRIMARY KEY CHECK (id = 1),
        trial_start_at TEXT NOT NULL,
        trial_end_at TEXT NOT NULL,
        last_seen_at TEXT NOT NULL,
        state TEXT NOT NULL
    )"""
    )

    cur.execute(
        """
    CREATE TABLE IF NOT EXISTS companies (
      id INTEGER PRIMARY KEY AUTOINCREMENT,
      name TEXT NOT NULL,
      match_key TEXT NOT NULL,
      whatsapp_phone TEXT NOT NULL,
      is_active INTEGER NOT NULL DEFAULT 1,
      created_at TEXT NOT NULL,

      auto_enabled INTEGER NOT NULL DEFAULT 0,
      next_send_at TEXT,
      last_send_at TEXT,
      last_send_status TEXT,
      last_error TEXT,
      last_send_files INTEGER NOT NULL DEFAULT 0
    )"""
    )

    ccols = _table_cols(cur, "companies")
    if "last_send_files" not in ccols:
        cur.execute("ALTER TABLE companies ADD COLUMN last_send_files INTEGER NOT NULL DEFAULT 0")

    cur.execute(
        """
    CREATE TABLE IF NOT EXISTS documents (
      id INTEGER PRIMARY KEY AUTOINCREMENT,
      source_path TEXT NOT NULL,
      filename TEXT NOT NULL,
      company_id INTEGER,
      status TEXT NOT NULL,            -- READY / UNMATCHED / ARCHIVED / MISSING
      size_bytes INTEGER NOT NULL DEFAULT 0,
      mtime REAL NOT NULL DEFAULT 0,
      created_at TEXT NOT NULL,
      updated_at TEXT NOT NULL,
      last_error TEXT
    )"""
    )

    dcols = _table_cols(cur, "documents")
    if "last_error" not in dcols:
        cur.execute("ALTER TABLE documents ADD COLUMN last_error TEXT")

    cur.execute(
        """
    CREATE TABLE IF NOT EXISTS daily_counters (
      day TEXT PRIMARY KEY,
      messages_sent INTEGER NOT NULL DEFAULT 0,
      files_sent INTEGER NOT NULL DEFAULT 0,
      bytes_sent INTEGER NOT NULL DEFAULT 0,
      updated_at TEXT NOT NULL
    )"""
    )

    cur.execute("CREATE INDEX IF NOT EXISTS idx_companies_active ON companies(is_active)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_docs_company_status ON documents(company_id,status)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_docs_status ON documents(status)")

    now = utcnow()
    cur.execute("SELECT trial_start_at, trial_end_at, state FROM license_state WHERE id=1")
    row = cur.fetchone()
    if not row:
        start = now
        end = now + timedelta(days=30)
        cur.execute(
            "INSERT INTO license_state (id, trial_start_at, trial_end_at, last_seen_at, state) VALUES (1, ?, ?, ?, ?)",
            (start.isoformat(), end.isoformat(), now.isoformat(), "TRIAL_ACTIVE"),
        )
    else:
        trial_end = datetime.fromisoformat(row[1])
        state = row[2]
        if state == "TRIAL_ACTIVE" and now > trial_end:
            cur.execute(
                "UPDATE license_state SET state=?, last_seen_at=? WHERE id=1",
                ("TRIAL_EXPIRED", now.isoformat()),
            )
        else:
            cur.execute("UPDATE license_state SET last_seen_at=? WHERE id=1", (now.isoformat(),))



    # --- Safe migrations ---
    ensure_table_columns("companies", {
    "created_at": "ALTER TABLE companies ADD COLUMN created_at TEXT",
    "is_active": "ALTER TABLE companies ADD COLUMN is_active INTEGER NOT NULL DEFAULT 1",
    "auto_enabled": "ALTER TABLE companies ADD COLUMN auto_enabled INTEGER NOT NULL DEFAULT 0",
    "next_send_at": "ALTER TABLE companies ADD COLUMN next_send_at TEXT",
    "last_send_at": "ALTER TABLE companies ADD COLUMN last_send_at TEXT",
    "last_send_status": "ALTER TABLE companies ADD COLUMN last_send_status TEXT",
    "last_error": "ALTER TABLE companies ADD COLUMN last_error TEXT",
    "last_send_files": "ALTER TABLE companies ADD COLUMN last_send_files INTEGER NOT NULL DEFAULT 0",
    "last_send_extra_count": "ALTER TABLE companies ADD COLUMN last_send_extra_count INTEGER NOT NULL DEFAULT 0",
    "last_send_status_text": "ALTER TABLE companies ADD COLUMN last_send_status_text TEXT",

    })
    ensure_table_columns("documents", {
    "source_path": "ALTER TABLE documents ADD COLUMN source_path TEXT",
    "size_bytes": "ALTER TABLE documents ADD COLUMN size_bytes INTEGER NOT NULL DEFAULT 0",
    "status": "ALTER TABLE documents ADD COLUMN status TEXT NOT NULL DEFAULT 'NEW'",
    "company_id": "ALTER TABLE documents ADD COLUMN company_id INTEGER",
    "updated_at": "ALTER TABLE documents ADD COLUMN updated_at TEXT",
    "last_error": "ALTER TABLE documents ADD COLUMN last_error TEXT",
    })

    con.commit()
    con.close()


def settings_get(key: str, default: str = "") -> str:
    con = db()
    cur = con.cursor()
    cur.execute("SELECT value FROM settings WHERE key=?", (key,))
    row = cur.fetchone()
    con.close()
    return row[0] if row else default


def settings_set(key: str, value: str) -> None:
    con = db()
    cur = con.cursor()
    cur.execute(
        "INSERT INTO settings(key,value) VALUES(?,?) ON CONFLICT(key) DO UPDATE SET value=excluded.value",
        (key, value),
    )
    con.commit()
    con.close()


def normalize_tr_phone(raw: str) -> str:
    s = (raw or "").strip().replace(" ", "").replace("-", "")
    if not s:
        return s
    if s.startswith("00"):
        s = "+" + s[2:]
    if s.startswith("0") and len(s) >= 10:
        if s.startswith("05"):
            return "+9" + s
        return "+90" + s[1:]
    if s.startswith("5") and len(s) == 10:
        return "+90" + s
    if s.startswith("+"):
        return s
    return s


def format_size(bytes_: int) -> str:
    if bytes_ < 1024:
        return f"{bytes_} B"
    kb = bytes_ / 1024
    if kb < 1024:
        return f"{int(kb)} KB"
    mb = kb / 1024
    if mb < 1024:
        return f"{mb:.1f} MB".replace(".0", "")
    gb = mb / 1024
    return f"{gb:.1f} GB".replace(".0", "")


def render_message_template(tpl: str, company_name: str, file_count: int) -> str:
    tpl = (tpl or "").strip()
    return (
        tpl.replace("%evraksayisi%", str(file_count))
        .replace("%firma%", company_name or "")
        .replace("%tarih%", local_now().strftime("%d.%m.%Y %H:%M"))
    )


def wa_health() -> Tuple[bool, str]:
    try:
        r = requests.get(WA_SENDER_BASE + "/health", timeout=5)
        if r.status_code == 200 and r.json().get("wa_ready") is True:
            return True, "OK"
        return False, (r.text[:250] or "Hazır değil")
    except Exception as e:
        return False, str(e)


def wa_send_batch(phone: str, message: str, file_paths: List[str], company_id: int = 0, company_name: str = "") -> tuple[bool, str]:
    """
    Node WA Sender'a toplu gönderim çağrısı.
    - Node /send-batch true döndürse bile results içindeki hataları kontrol eder.
    - En az 1 adet (text veya media) başarı olmadan OK saymaz.
    """
    try:
        import requests

        # önce node health: wa_ready mı?
        try:
            hr = requests.get("http://127.0.0.1:3210/health", timeout=5)
            hj = hr.json() if hr.content else {}
            if not hj.get("wa_ready"):
                return False, "WhatsApp servisi hazır değil (node wa_ready=false)."
        except Exception:
            # health okunamazsa devam et, send-batch cevapları belirleyici olsun
            pass

        per_min = int(settings_get("per_file_min_s", "8") or 8) * 1000
        per_max = int(settings_get("per_file_max_s", "15") or 15) * 1000

        payload = {
            "to": phone,
            "message": message,
            "company_id": int(company_id) if company_id else 0,
            "company_name": (company_name or ""),
            "perFileMinMs": per_min,
            "perFileMaxMs": per_max,
            "files": [{"path": p, "caption": ""} for p in (file_paths or [])],
        }

        r = requests.post("http://127.0.0.1:3210/send-batch", json=payload, timeout=240)
        j = r.json() if r.content else {}

        if not j.get("ok"):
            return False, j.get("error") or f"WA sender error: {str(j)[:240]}"

        results = j.get("results") or []
        if not isinstance(results, list):
            results = []

        ok_count = 0
        fail_msgs = []
        for it in results:
            try:
                it_ok = bool(it.get("ok", True))  # text item'larda ok alanı olmayabilir
                mid = (it.get("message_id") or "") if isinstance(it, dict) else ""
                it_type = (it.get("type") or "") if isinstance(it, dict) else ""
                it_err = (it.get("error") or "") if isinstance(it, dict) else ""
                if it_ok and (mid or it_type == "text"):
                    ok_count += 1
                else:
                    if it_err:
                        fail_msgs.append(f"{it_type}:{it_err}")
                    else:
                        # message_id yoksa da hata say
                        if it_type:
                            fail_msgs.append(f"{it_type}:NO_MESSAGE_ID")
                        else:
                            fail_msgs.append("ITEM_FAIL")
            except Exception:
                fail_msgs.append("ITEM_PARSE_FAIL")

        # Hiç sonuç yoksa (message+files boş gibi) hata
        if ok_count == 0:
            if fail_msgs:
                return False, "Gönderim başarısız: " + " | ".join(fail_msgs)[:240]
            return False, "Gönderim başarısız (sonuç boş veya message_id yok)."

        # kısmi hata varsa yine de hata dönelim ki main.py DB'de ERROR bıraksın
        if fail_msgs:
            return False, "Kısmi hata: " + " | ".join(fail_msgs)[:240]

        return True, "OK"
    except Exception as e:
        return False, str(e)


def get_trial_info() -> dict:
    con = db()
    cur = con.cursor()
    cur.execute("SELECT trial_start_at, trial_end_at, last_seen_at, state FROM license_state WHERE id=1")
    row = cur.fetchone()
    con.close()
    if not row:
        return {"app": APP_NAME, "state": "UNKNOWN", "remaining_days": 0}
    start_at, end_at, last_seen_at, state = row
    now = utcnow()
    end_dt = datetime.fromisoformat(end_at)
    remaining_days = max(0, int((end_dt - now).total_seconds() // 86400))
    return {
        "app": APP_NAME,
        "state": state,
        "trial_start_at": start_at,
        "trial_end_at": end_at,
        "last_seen_at": last_seen_at,
        "remaining_days": remaining_days,
    }


def companies_all() -> List[dict]:
    con = db()
    cur = con.cursor()
    cur.execute(
        """
      SELECT id, name, match_key, whatsapp_phone, is_active, created_at,
             auto_enabled, next_send_at, last_send_at, last_send_status, last_error, last_send_files
      FROM companies
      ORDER BY id DESC
    """
    )
    rows = cur.fetchall()
    con.close()
    out: List[dict] = []
    for r in rows:
        out.append(
            {
                "id": r[0],
                "name": r[1],
                "match_key": r[2],
                "whatsapp_phone": r[3],
                "is_active": int(r[4] or 0),
                "created_at": r[5],
                "auto_enabled": int(r[6] or 0),
                "next_send_at": r[7] or "",
                "last_send_at": r[8] or "",
                "last_send_status": r[9] or "",
                "last_error": r[10] or "",
                "last_send_files": int(r[11] or 0),
            }
        )
    return out


def company_by_id(cid: int) -> Optional[dict]:
    con = db()
    cur = con.cursor()
    cur.execute(
        """
      SELECT id, name, match_key, whatsapp_phone, is_active, created_at,
             auto_enabled, next_send_at, last_send_at, last_send_status, last_error, last_send_files
      FROM companies WHERE id=?
    """,
        (cid,),
    )
    r = cur.fetchone()
    con.close()
    if not r:
        return None
    return {
        "id": r[0],
        "name": r[1],
        "match_key": r[2],
        "whatsapp_phone": r[3],
        "is_active": int(r[4] or 0),
        "created_at": r[5],
        "auto_enabled": int(r[6] or 0),
        "next_send_at": r[7] or "",
        "last_send_at": r[8] or "",
        "last_send_status": r[9] or "",
        "last_error": r[10] or "",
        "last_send_files": int(r[11] or 0),
    }


def ready_stats_by_company() -> Dict[int, dict]:
    con = db()
    cur = con.cursor()
    cur.execute(
        """
      SELECT company_id, COUNT(*), COALESCE(SUM(size_bytes),0)
      FROM documents
      WHERE status='READY' AND company_id IS NOT NULL
      GROUP BY company_id
    """
    )
    rows = cur.fetchall()
    con.close()
    stats: Dict[int, dict] = {}
    for cid, cnt, total_bytes in rows:
        stats[int(cid)] = {"ready_count": int(cnt), "ready_bytes": int(total_bytes)}
    return stats


def scan_source_folder() -> Tuple[int, int, int, str]:
    source_dir = settings_get("source_dir", "").strip()
    if not source_dir:
        return 0, 0, 0, "Kaynak klasör ayarlı değil."
    p = Path(source_dir)
    if not p.exists() or not p.is_dir():
        return 0, 0, 0, f"Kaynak klasör bulunamadı: {source_dir}"

    comps = [c for c in companies_all() if c["is_active"] == 1]
    exts = {".pdf", ".jpg", ".jpeg", ".png", ".xls", ".xlsx", ".zip"}

    scanned = matched = inserted = 0
    con = db()
    cur = con.cursor()
    now = utcnow().isoformat()

    for fp in p.iterdir():
        if not fp.is_file():
            continue
        if fp.suffix.lower() not in exts:
            continue
        scanned += 1

        filename = fp.name
        st = fp.stat()
        size_bytes = st.st_size
        mtime = st.st_mtime

        best = None
        best_len = -1
        for c in comps:
            mk = (c["match_key"] or "").strip()
            if mk and mk.lower() in filename.lower():
                if len(mk) > best_len:
                    best = c
                    best_len = len(mk)

        company_id = best["id"] if best else None
        status = "READY" if company_id else "UNMATCHED"

        cur.execute("SELECT id, mtime FROM documents WHERE source_path=?", (str(fp),))
        row = cur.fetchone()
        if not row:
            cur.execute(
                "INSERT INTO documents(source_path, filename, company_id, status, size_bytes, mtime, created_at, updated_at, last_error) VALUES(?,?,?,?,?,?,?,?,?)",
                (str(fp), filename, company_id, status, int(size_bytes), float(mtime), now, now, ""),
            )
            inserted += 1
        else:
            doc_id, old_mtime = row
            if float(old_mtime) != float(mtime):
                cur.execute(
                    "UPDATE documents SET filename=?, company_id=?, status=?, size_bytes=?, mtime=?, updated_at=?, last_error=? WHERE id=?",
                    (filename, company_id, status, int(size_bytes), float(mtime), now, "", doc_id),
                )

        if company_id:
            matched += 1

    con.commit()
    con.close()
    return scanned, matched, inserted, ""



def cleanup_missing_ready() -> int:
    """READY görünüp diskte bulunmayan kayıtları MISSING'e çevirir.
    UI sayaçlarının şişmesini engeller."""
    con = db()
    cur = con.cursor()
    cur.execute("SELECT id, source_path FROM documents WHERE status='READY'")
    rows = cur.fetchall()
    miss = []
    now = utcnow().isoformat()
    for doc_id, p in rows:
        if not os.path.exists(p):
            miss.append((now, int(doc_id)))
    if miss:
        cur.executemany(
            "UPDATE documents SET status='MISSING', last_error='Dosya bulunamadı (diskte yok).', updated_at=? WHERE id=?",
            miss,
        )
    con.commit()
    con.close()
    return len(miss)


def documents_ready_rows(cid: int) -> List[tuple]:
    con = db()
    cur = con.cursor()
    cur.execute("SELECT id, source_path FROM documents WHERE company_id=? AND status='READY' ORDER BY id ASC", (cid,))
    rows = cur.fetchall()
    con.close()
    return rows

from collections import deque

# Eğer dosyada yoksa:
# RECENT_JOBS = deque(maxlen=200)
# RECENT_TTL_SEC = 5

def _fmt_tr_dt(s: str) -> str:
    """ISO/UTC/naive -> TR 'dd.mm.yyyy HH:MM'"""
    s = (s or "").strip()
    if not s:
        return ""
    try:
        dt = parse_iso(s)
        if dt:
            return dt.astimezone(TR_TZ).strftime("%d.%m.%Y %H:%M")
    except Exception:
        pass
    try:
        s2 = s.replace("T", " ").strip()[:16]
        dt2 = datetime.strptime(s2, "%Y-%m-%d %H:%M").replace(tzinfo=TR_TZ)
        return dt2.strftime("%d.%m.%Y %H:%M")
    except Exception:
        return s.replace("T", " ")[:16]


def push_recent_job_done(company_id: int, company_name: str, total_files: int):
    """Gönderim bittiğinde 5 sn sidebar/job-tracking'de görünsün."""
    try:
        RECENT_JOBS.appendleft({
            "type": "done",
            "id": int(company_id),
            "name": company_name or "-",
            "x": int(total_files),
            "y": int(total_files),
            "when": iso(utcnow()),
            "when_tr": _fmt_tr_dt(iso(utcnow())),
            "ts": time.time(),
        })
    except Exception:
        pass

def archive_ready_docs(company: dict) -> Tuple[int, str]:
    """READY evrakları arşive taşır.

    Arşiv yapısı (aylık klasör):
      <archive_dir>/<YYYY-MM>/<Firma Adı>/dosyalar...
    """
    archive_dir = settings_get("archive_dir", "").strip()
    if not archive_dir:
        return 0, "Arşiv klasörü ayarlı değil."
    ap = Path(archive_dir)
    ap.mkdir(parents=True, exist_ok=True)

    safe_name = (
        "".join([c if c.isalnum() or c in " _-" else "_" for c in company["name"]]).strip()
        or f"firma_{company['id']}"
    )

    # Aylık klasör (TR saatine göre)
    try:
        now_tr = datetime.now(TR_TZ)
    except Exception:
        now_tr = datetime.now()
    sub = f"{now_tr.year}-{now_tr.month:02d}"

    # <archive_dir>/<YYYY-MM>/<Firma>/
    target_dir = ap / sub / safe_name
    target_dir.mkdir(parents=True, exist_ok=True)

    con = db()
    cur = con.cursor()
    cur.execute(
        "SELECT id, source_path FROM documents WHERE company_id=? AND status='READY' ORDER BY id ASC",
        (company["id"],),
    )
    docs = cur.fetchall()

    moved = 0
    now = utcnow().isoformat()

    for doc_id, src_path in docs:
        src = Path(src_path)
        if not src.exists():
            cur.execute(
                "UPDATE documents SET status='MISSING', last_error=?, updated_at=? WHERE id=?",
                ("Dosya bulunamadı (diskte yok).", now, doc_id),
            )
            continue
        dst = target_dir / src.name
        # isim çakışması olursa benzersizleştir
        if dst.exists():
            stem = dst.stem
            suf = dst.suffix
            n = 1
            while True:
                cand = target_dir / f"{stem} ({n}){suf}"
                if not cand.exists():
                    dst = cand
                    break
                n += 1
        try:
            shutil.move(str(src), str(dst))
            moved += 1
            cur.execute("UPDATE documents SET status='ARCHIVED', updated_at=? WHERE id=?", (now, doc_id))
        except Exception as e:
            cur.execute(
                "UPDATE documents SET last_error=?, updated_at=? WHERE id=?",
                (str(e)[:300], now, doc_id),
            )

    con.commit()
    con.close()
    return moved, ""

def get_today_key() -> str:
    return local_now().strftime("%Y-%m-%d")


def daily_get() -> dict:
    day = get_today_key()
    con = db()
    cur = con.cursor()
    cur.execute("SELECT day, messages_sent, files_sent, bytes_sent, updated_at FROM daily_counters WHERE day=?", (day,))
    row = cur.fetchone()
    if not row:
        now = utcnow().isoformat()
        cur.execute(
            "INSERT INTO daily_counters(day,messages_sent,files_sent,bytes_sent,updated_at) VALUES(?,?,?,?,?)",
            (day, 0, 0, 0, now),
        )
        con.commit()
        row = (day, 0, 0, 0, now)
    con.close()
    return {"day": row[0], "messages_sent": row[1], "files_sent": row[2], "bytes_sent": row[3], "updated_at": row[4]}


def daily_add(messages: int, files: int, bytes_: int) -> None:
    day = get_today_key()
    con = db()
    cur = con.cursor()
    now = utcnow().isoformat()
    cur.execute(
        """
      INSERT INTO daily_counters(day,messages_sent,files_sent,bytes_sent,updated_at)
      VALUES(?,?,?,?,?)
      ON CONFLICT(day) DO UPDATE SET
        messages_sent=messages_sent+excluded.messages_sent,
        files_sent=files_sent+excluded.files_sent,
        bytes_sent=bytes_sent+excluded.bytes_sent,
        updated_at=excluded.updated_at
    """,
        (day, messages, files, bytes_, now),
    )
    con.commit()
    con.close()


def limits() -> dict:
    return {
        "daily_msg_limit": int(settings_get("daily_msg_limit", "1100") or "1100"),
        "daily_file_limit": int(settings_get("daily_file_limit", "1100") or "1100"),
        "daily_mb_limit": int(settings_get("daily_mb_limit", "1500") or "1500"),
    }


_autoscan_lock = threading.Lock()
_autoscan_pending = False
_last_scan_ts = 0.0
_observer: Observer | None = None



def set_last_scan_now() -> None:
    global _last_scan_ts
    _last_scan_ts = time.time()


def last_scan_text() -> str:
    if _last_scan_ts <= 0:
        return "-"
    return datetime.fromtimestamp(_last_scan_ts).strftime("%Y-%m-%d %H:%M:%S")


def schedule_autoscan() -> None:
    global _autoscan_pending
    with _autoscan_lock:
        _autoscan_pending = True


def autoscan_loop() -> None:
    global _autoscan_pending
    while True:
        time.sleep(0.5)
        do_scan = False
        with _autoscan_lock:
            if _autoscan_pending:
                do_scan = True
                _autoscan_pending = False
        if do_scan:
            time.sleep(1.5)
            try:
                scanned, matched, inserted, err = scan_source_folder()
                set_last_scan_now()
                cleanup_missing_ready()
                if err:
                    settings_set("last_scan_err", err)
                    settings_set("last_scan_msg", "")
                else:
                    settings_set("last_scan_err", "")
                    settings_set("last_scan_msg", f"Tarandı:{scanned} | Eşleşen:{matched} | Yeni:{inserted}")
            except Exception as e:
                settings_set("last_scan_err", str(e))
                settings_set("last_scan_msg", "")


class _Handler(FileSystemEventHandler):
    def on_any_event(self, event):
        if getattr(event, "is_directory", False):
            return
        schedule_autoscan()

def _mark_company_sent_with_extra(
    company_id: int,
    existing_paths: list[str],
    status: str,
    aerr: str,
    extra_label: str = "Ek",
) -> tuple[int, str]:
    """Gönderim sonrası company kaydını günceller ve ek sayacını artırır.

    - last_send_extra_count: her başarılı gönderimde +1
    - last_send_status_text: "Mesaj Gönderildi · Ek N" gibi kullanıcı metni
    """
    con4 = db()
    cur4 = con4.cursor()

    # önce mevcut sayacı çek
    try:
        cur4.execute(
            "SELECT COALESCE(last_send_extra_count,0) FROM companies WHERE id=?",
            (int(company_id),),
        )
        extra_n = int((cur4.fetchone() or [0])[0]) + 1
    except Exception:
        extra_n = 1

    status_text = f"Mesaj Gönderildi · {extra_label} {extra_n}"

    # tek update ile yaz
    cur4.execute(
        """
        UPDATE companies
        SET last_send_at=?,
            last_send_status=?,
            last_error=?,
            last_send_files=?,
            last_send_extra_count=?,
            last_send_status_text=?
        WHERE id=?
        """,
        (iso(utcnow()), status, aerr or "", int(len(existing_paths)), int(extra_n), status_text, int(company_id)),
    )

    con4.commit()
    con4.close()
    return extra_n, status_text

def start_watchdog() -> None:
    global _observer
    source_dir = settings_get("source_dir", "").strip()
    if not source_dir:
        return
    p = Path(source_dir)
    if not p.exists() or not p.is_dir():
        return
    if _observer is not None:
        return
    threading.Thread(target=autoscan_loop, daemon=True).start()
    _observer = Observer()
    _observer.schedule(_Handler(), str(p), recursive=False)
    _observer.daemon = True
    _observer.start()


def watchdog_enabled() -> bool:
    return _observer is not None


def send_status_line() -> str:
    base = f"Otomatik izleme: {'Açık' if watchdog_enabled() else 'Kapalı'} | Son tarama: {last_scan_text()}"
    last_msg = settings_get("last_scan_msg", "").strip()
    if last_msg:
        base += f" | {last_msg}"
    return base


_scheduler_thread_started = False

def scheduler_tick() -> None:
    while True:
        time.sleep(20)

        try:
            ok, info = wa_health()
        except Exception as e:
            try:
                err("SCHED_WA_HEALTH_ERR", e)
            except Exception:
                pass
            continue

        if not ok:
            try:
                dbg("SCHED_SKIP_WA_NOT_READY", info=info)
            except Exception:
                pass
            continue

        now_utc = utcnow()

        # 1) Aday işleri çek
        try:
            con = db()
            cur = con.cursor()
            cur.execute(
                """
                SELECT id, next_send_at
                FROM companies
                WHERE is_active=1 AND (auto_enabled=1 OR COALESCE(last_send_status,'')='QUEUED')
                  AND next_send_at IS NOT NULL AND next_send_at != ''
                """
            )
            rows = cur.fetchall()
            con.close()
        except Exception as e:
            try:
                err("SCHED_DB_READ_ERR", e)
            except Exception:
                pass
            continue

        try:
            if rows:
                dbg("SCHED_TICK", candidates=len(rows))
        except Exception:
            pass

        for cid, next_send_at_str in rows:
            cid = int(cid)
            next_send_at_str = (next_send_at_str or "").strip()

            # 2) next_send_at parse + zaman kontrol
            try:
                next_send_at = parse_iso(next_send_at_str)
            except Exception:
                next_send_at = None

            if not next_send_at:
                try:
                    warn("SCHED_SKIP_BAD_NEXT_SEND_AT", company_id=cid, next_send_at=next_send_at_str)
                except Exception:
                    pass
                continue

            if next_send_at > now_utc:
                # henüz zamanı değil
                continue

            comp = None
            try:
                comp = company_by_id(cid)
            except Exception as e:
                try:
                    err("SCHED_COMPANY_BY_ID_ERR", e, company_id=cid)
                except Exception:
                    pass

            if not comp:
                continue

            # 3) Ready docs kontrol
            try:
                ready_rows = documents_ready_rows(cid)
            except Exception as e:
                try:
                    err("SCHED_READY_ROWS_ERR", e, company_id=cid)
                except Exception:
                    pass
                continue

            existing_paths: List[str] = []
            missing_ids: List[int] = []
            for doc_id, path in ready_rows:
                p = str(path or "")
                if p and Path(p).exists():
                    existing_paths.append(p)
                else:
                    if doc_id is not None:
                        try:
                            missing_ids.append(int(doc_id))
                        except Exception:
                            pass

            if missing_ids:
                try:
                    conx = db()
                    cx = conx.cursor()
                    now_iso = utcnow().isoformat()
                    cx.executemany(
                        "UPDATE documents SET status='MISSING', last_error=?, updated_at=? WHERE id=?",
                        [("Dosya bulunamadı (diskte yok).", now_iso, did) for did in missing_ids],
                    )
                    conx.commit()
                    conx.close()
                    warn("SCHED_DOCS_MISSING", company_id=cid, missing=len(missing_ids))
                except Exception as e:
                    try:
                        err("SCHED_MARK_MISSING_ERR", e, company_id=cid)
                    except Exception:
                        pass

            if not existing_paths:
                # dosya yoksa +1 saat ertele
                try:
                    con2 = db()
                    c2 = con2.cursor()
                    c2.execute(
                        "UPDATE companies SET last_send_status=?, last_error=?, last_send_files=?, next_send_at=? WHERE id=?",
                        ("NO_FILES", "Hazır dosya yok / dosyalar kayıp.", 0, iso(now_utc + timedelta(hours=1)), cid),
                    )
                    con2.commit()
                    con2.close()
                    warn("SCHED_NO_FILES_POSTPONE", company_id=cid)
                except Exception as e:
                    try:
                        err("SCHED_NO_FILES_UPDATE_ERR", e, company_id=cid)
                    except Exception:
                        pass
                continue

            # 4) Limit kontrol
            try:
                d = daily_get()
                lim = limits()
                planned_files = len(existing_paths)
                planned_msgs = 1 + planned_files
                planned_mb = int(sum(Path(p).stat().st_size for p in existing_paths) / (1024 * 1024))
                used_mb = int(d["bytes_sent"] / (1024 * 1024))
            except Exception as e:
                try:
                    err("SCHED_LIMIT_CALC_ERR", e, company_id=cid)
                except Exception:
                    pass
                continue

            if (
                d["messages_sent"] + planned_msgs > lim["daily_msg_limit"]
                or d["files_sent"] + planned_files > lim["daily_file_limit"]
                or used_mb + planned_mb > lim["daily_mb_limit"]
            ):
                try:
                    con2 = db()
                    c2 = con2.cursor()
                    c2.execute(
                        "UPDATE companies SET last_send_status=?, last_error=? WHERE id=?",
                        ("PAUSED_LIMIT", "Günlük limit doldu", cid),
                    )
                    con2.commit()
                    con2.close()
                    warn("SCHED_PAUSED_LIMIT", company_id=cid, planned_files=planned_files, planned_mb=planned_mb)
                except Exception as e:
                    try:
                        err("SCHED_PAUSED_LIMIT_ERR", e, company_id=cid)
                    except Exception:
                        pass
                continue

            # 5) WhatsApp gönder (tek gönderim kilidi)
            acquired = SEND_MUTEX.acquire(blocking=False)
            if not acquired:
                dbg("SCHED_BUSY_SKIP", company_id=cid)
                break
            try:
                phone = normalize_tr_phone(comp["whatsapp_phone"])
                msg_tpl = settings_get("send_message", "Vergi/SSK Tahakkuk evraklarınız ektedir.")
                msg = render_message_template(msg_tpl, comp.get("name", ""), len(existing_paths))

                con2 = db()
                c2 = con2.cursor()
                c2.execute("UPDATE companies SET last_send_status=?, last_error=? WHERE id=?", ("SENDING", "", cid))
                con2.commit()
                con2.close()

                dbg("SCHED_SENDING", company_id=cid, files=len(existing_paths))

                ok2, info2 = wa_send_batch(phone, msg, existing_paths, company_id=cid, company_name=comp.get("name",""))
                if not ok2:
                    con3 = db()
                    c3 = con3.cursor()
                    c3.execute("UPDATE companies SET last_send_status=?, last_error=? WHERE id=?", ("ERROR", info2, cid))
                    con3.commit()
                    con3.close()
                    continue

                moved, aerr = archive_ready_docs(comp)
                status = "SENT" if not aerr else "SENT_ARCHIVE_ERROR"
                err_txt = aerr or ""

                total_bytes = 0
                for p in existing_paths:
                    try:
                        total_bytes += Path(p).stat().st_size
                    except Exception:
                        pass
                daily_add(messages=1 + len(existing_paths), files=len(existing_paths), bytes_=total_bytes)

                # ✅ başarılı: şirket kaydı + Ek sayacı + plan temizliği
                _mark_company_sent_with_extra(int(comp["id"]), existing_paths, status, aerr, company_name=comp.get("name",""))
                push_recent_job_done(company_id=int(comp["id"]), company_name=comp.get("name",""), files=len(existing_paths))

                con4 = db()
                c4 = con4.cursor()
                c4.execute(
                    "UPDATE companies SET last_send_status=?, last_error=? WHERE id=?",
                    (status, err_txt, cid),
                )
                con4.commit()
                con4.close()

            finally:
                try:
                    SEND_MUTEX.release()
                except Exception:
                    pass
def build_ctx(active: str, page_title: str) -> dict:
    trial = get_trial_info()
    wa_ok, wa_info = wa_health()
    d = daily_get()
    lim = limits()
    used_mb = int(d["bytes_sent"] / (1024 * 1024))
    remaining = {
        "msg": max(0, lim["daily_msg_limit"] - d["messages_sent"]),
        "file": max(0, lim["daily_file_limit"] - d["files_sent"]),
        "mb": max(0, lim["daily_mb_limit"] - used_mb),
    }

    # İş Takibi TTL (varsayılan 24 saat)
    try:
        job_done_ttl_sec = int(settings_get("job_done_ttl_sec", "86400") or 86400)
    except Exception:
        job_done_ttl_sec = 86400
    if job_done_ttl_sec < 60:
        job_done_ttl_sec = 60
    job_done_ttl_hours = int(job_done_ttl_sec // 3600) if job_done_ttl_sec >= 3600 else 1

    return {
        "title": APP_NAME,
        "page_title": page_title,
        "active": active,
        "trial": trial,
        "wa_ok": wa_ok,
        "wa_info": wa_info,
        "daily": d,
        "limits": lim,
        "remaining": remaining,
        "job_done_ttl_hours": job_done_ttl_hours,
        "job_done_ttl_sec": job_done_ttl_sec,
    }

app = FastAPI(title=APP_NAME)
templates = Jinja2Templates(directory="app/templates")
app.mount("/static", StaticFiles(directory="app/static"), name="static")

# === STATIC & TEMPLATES PATH FIX (PyInstaller uyumlu) ===
import sys
from pathlib import Path
from fastapi.staticfiles import StaticFiles

def _resource_dir(*parts: str) -> Path:
    """
    Dev ortamında: proje/app altından okur
    PyInstaller onefile'da: sys._MEIPASS altından okur
    """
    if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
        base = Path(sys._MEIPASS)  # PyInstaller extract dir
    else:
        # main.py genelde app/main.py ise:
        base = Path(__file__).resolve().parent
    return (base.joinpath(*parts)).resolve()

# Sen pyinstaller'da --add-data ile şunları kopyalıyoruz:
# templates -> <base>/templates
# static    -> <base>/static
STATIC_DIR = _resource_dir("static")
TEMPLATES_DIR = _resource_dir("templates")

# Static klasörü yoksa fallback: app/static arayan eski yapıyı da tolere et
if not STATIC_DIR.exists():
    # bazı paketlerde static farklı konumlanmış olabilir
    STATIC_DIR = _resource_dir("app", "static")

app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")
# === /STATIC & TEMPLATES PATH FIX ===


# ===============================
# WhatsApp ACK (DELIVERED/READ) webhook
# Node wa_sender/index.js buraya POST atar
# ===============================
from fastapi import Body
from fastapi.responses import JSONResponse

from fastapi.responses import Response

@app.get("/favicon.ico")
def favicon():
    return Response(status_code=204)

def _ensure_company_ack_columns():
    """companies tablosunda ack alanları yoksa ekle (kayıpsız)."""
    try:
        con = db()
        cur = con.cursor()
        cur.execute("PRAGMA table_info(companies)")
        cols = {r[1] for r in cur.fetchall()}  # (cid, name, type,...)

        alters = []
        if "last_ack" not in cols:
            alters.append("ALTER TABLE companies ADD COLUMN last_ack TEXT")
        if "last_ack_at" not in cols:
            alters.append("ALTER TABLE companies ADD COLUMN last_ack_at TEXT")
        if "last_ack_message_id" not in cols:
            alters.append("ALTER TABLE companies ADD COLUMN last_ack_message_id TEXT")
        if "last_ack_to" not in cols:
            alters.append("ALTER TABLE companies ADD COLUMN last_ack_to TEXT")

        for sql in alters:
            cur.execute(sql)

        con.commit()
        con.close()
    except Exception:
        # tablo kilitli vs. olabilir, sessiz geç
        try:
            con.close()
        except Exception:
            pass


@app.post("/wa/ack")
async def wa_ack(request: Request):
    """
    Node -> FastAPI ACK endpoint.
    - company_id / to boş gelse bile logları bozmaz
    - to varsa firmayı DB'den tahmin etmeye çalışır
    - event/ack bilgilerini güvenli şekilde yazar
    """
    try:
        data = await request.json()
    except Exception:
        data = {}

    event = str(data.get("event") or "").strip()          # "sent" | "delivered" | "read" | "error" ...
    message_id = str(data.get("message_id") or "").strip()
    ack = int(data.get("ack") or 0)
    ack_text = str(data.get("ack_text") or "").strip()
    to_raw = str(data.get("to") or "").strip()
    company_id = int(data.get("company_id") or 0)

    # Normalize phone if possible (tolerant)
    to_norm = ""
    try:
        to_norm = normalize_tr_phone(to_raw) if to_raw else ""
    except Exception:
        to_norm = to_raw

    # company_id boşsa: to üzerinden firmayı tahmin et
    inferred_company_id = 0
    inferred_company_name = ""
    if company_id <= 0 and to_norm:
        try:
            con = db()
            cur = con.cursor()
            # whatsapp_phone alanı +90'lı / 0'lı / boşluklu olabilir -> kaba eşleşme
            # (normalize edilmiş eşleşmeyi garanti etmek için istersen DB'de normalize kolon tutarız)
            cur.execute("SELECT id, name, whatsapp_phone FROM companies")
            rows = cur.fetchall()
            con.close()

            for cid, cname, phone in rows:
                try:
                    if normalize_tr_phone(phone) == to_norm:
                        inferred_company_id = int(cid)
                        inferred_company_name = str(cname or "")
                        break
                except Exception:
                    continue
        except Exception as e:
            warn("WA_ACK_INFER_FAIL", err=str(e))

    final_company_id = company_id if company_id > 0 else inferred_company_id
    final_company_name = inferred_company_name

    # Log: boş bile olsa stabil
    dbg(
        "WA_ACK_IN",
        event=event,
        message_id=message_id,
        ack=ack,
        ack_text=ack_text,
        to=to_norm or to_raw,
        company_id=final_company_id,
        company_name=final_company_name,
    )

    # İstersen DB event log'a yaz (company_id varsa)
    if final_company_id > 0:
        try:
            # status alanını sadeleştir
            st = "ACK"
            if event:
                st = f"ACK_{event.upper()}"
            msg = f"ACK {event or '-'} | ack={ack} {ack_text or ''} | mid={message_id or '-'}"
            db_log(final_company_id, final_company_name, st, msg, 0)
        except Exception:
            pass

    return JSONResponse({"ok": True, "company_id": final_company_id})

# ===============================
# AUTO START WA SENDER (Node)
# ===============================
import os
import sys
import subprocess
from pathlib import Path
from datetime import datetime

def _is_frozen() -> bool:
    return getattr(sys, "frozen", False) is True

def _base_dir() -> Path:
    # EXE çalışıyorsa: exe'nin bulunduğu klasör
    if _is_frozen():
        return Path(sys.executable).resolve().parent
    # Source çalışıyorsa: app/main.py -> proje kökü
    return Path(__file__).resolve().parent.parent

def _logs_dir() -> Path:
    d = _base_dir() / "logs"
    d.mkdir(parents=True, exist_ok=True)
    return d

def _wa_sender_dir() -> Path:
    """wa_sender dizinini bul (EXE/onefile uyumlu).

    Öncelik:
    1) EXE'nin bulunduğu dizin / wa_sender  (portable kullanım)
    2) PyInstaller onefile extract dizini (sys._MEIPASS) / wa_sender  (add-data ile paketlenmişse)
    3) Geliştirme ortamında proje kökü / wa_sender
    """
    # 1) EXE yanında
    cand = _base_dir() / "wa_sender"
    if cand.exists():
        return cand

    # 2) PyInstaller onefile (_MEIPASS)
    try:
        import sys as _sys
        if getattr(_sys, "frozen", False) and hasattr(_sys, "_MEIPASS"):
            meipass = Path(getattr(_sys, "_MEIPASS"))
            cand2 = (meipass / "wa_sender").resolve()
            if cand2.exists():
                return cand2
    except Exception:
        pass

    # 3) Dev ortamı: repo kökü altında aransın
    try:
        repo_root = Path(__file__).resolve().parents[1]  # app/main.py -> repo/
        cand3 = (repo_root / "wa_sender").resolve()
        if cand3.exists():
            return cand3
    except Exception:
        pass

    return cand  # default (yoksa da log'a düşer)

def _probe_wa_sender() -> bool:
    # wa_sender /health endpoint'i (3210)
    try:
        import requests
        r = requests.get("http://127.0.0.1:3210/health", timeout=1.2)
        if r.ok:
            j = r.json()
            return bool(j.get("ok")) and bool(j.get("wa_ready"))
    except Exception:
        return False
    return False

def start_wa_sender_if_needed() -> None:
    """WA Sender çalışmıyorsa Node process'i başlat."""
    try:
        # Zaten hazırsa hiçbir şey yapma
        if _probe_wa_sender():
            try:
                dbg("WA_SENDER_ALREADY_RUNNING")
            except Exception:
                pass
            return

        wa_dir = _wa_sender_dir()
        if not wa_dir.exists():
            # release paketinde wa_sender klasörü exe'nin yanında olmalı
            try:
                err("WA_SENDER_DIR_NOT_FOUND", None, wa_dir=str(wa_dir))
            except Exception:
                print("WA sender klasörü bulunamadı:", wa_dir)
            return

        index_js = wa_dir / "index.js"
        if not index_js.exists():
            try:
                err("WA_SENDER_INDEX_NOT_FOUND", None, index_js=str(index_js))
            except Exception:
                print("WA sender index.js bulunamadı:", index_js)
            return

        logf = _logs_dir() / "wa_sender.log"
        env = os.environ.copy()
        env["FASTAPI_BASE"] = "http://127.0.0.1:8787"
        env["ACK_WEBHOOK"] = "http://127.0.0.1:8787/wa/ack"

        # Windows: ayrı process, paneli bloklamasın
        creationflags = 0
        if os.name == "nt":
            creationflags = subprocess.CREATE_NEW_PROCESS_GROUP | subprocess.DETACHED_PROCESS

        with open(logf, "a", encoding="utf-8") as f:
            f.write(f"\n[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] START_WA_SENDER\n")
            f.flush()

        p = subprocess.Popen(
            ["node", "index.js"],
            cwd=str(wa_dir),
            env=env,
            stdout=open(logf, "a", encoding="utf-8"),
            stderr=open(logf, "a", encoding="utf-8"),
            creationflags=creationflags,
        )

        try:
            dbg("WA_SENDER_STARTED", pid=getattr(p, "pid", None), cwd=str(wa_dir))
        except Exception:
            print("WA Sender başlatıldı, PID:", getattr(p, "pid", None))

    except Exception as e:
        try:
            err("WA_SENDER_START_FAIL", e)
        except Exception:
            print("WA Sender start fail:", e)

# ✅ Startup: panel kalkınca WA Sender da kalksın
@app.on_event("startup")
def _startup_autostart_wa_sender():
    start_wa_sender_if_needed()


from fastapi import Request
from fastapi.responses import HTMLResponse

@app.get("/", response_class=HTMLResponse)
def dashboard_page(request: Request):
    """
    Premium Dashboard KPI'ları:
      - kpi_pending: sadece GELECEK planlı işler (next_send_at >= now)
      - kpi_companies_active: aktif firma sayısı
      - kpi_errors_24h: son 24 saat hata sayısı (companies üzerinden)
      - last_error_preview: tek satır son hata (opsiyonel)
    """
    ctx = build_ctx("dashboard", "Dashboard")

    con = db()
    cur = con.cursor()

    # 1) Aktif firma sayısı
    try:
        cur.execute("SELECT COUNT(1) FROM companies WHERE COALESCE(is_active,1)=1")
        kpi_companies_active = int((cur.fetchone() or [0])[0])
    except Exception:
        kpi_companies_active = 0

    # 2) Planlı / bekleyen iş sayısı (SADECE GELECEK)
    # next_send_at ISO olabilir: "2026-03-03T04:00:00+00:00" -> ilk 19 + T->' '
    try:
        cur.execute(
            """
            SELECT COUNT(1)
            FROM companies
            WHERE COALESCE(next_send_at,'') <> ''
              AND datetime(REPLACE(substr(next_send_at,1,19),'T',' ')) >= datetime('now')
            """
        )
        kpi_pending = int((cur.fetchone() or [0])[0])
    except Exception:
        kpi_pending = 0

    # 3) Son 24 saat hata sayısı
    try:
        cur.execute(
            """
            SELECT COUNT(1)
            FROM companies
            WHERE COALESCE(last_send_status,'')='ERROR'
              AND COALESCE(last_send_at,'') <> ''
              AND datetime(REPLACE(substr(last_send_at,1,19),'T',' ')) >= datetime('now','-1 day')
            """
        )
        kpi_errors_24h = int((cur.fetchone() or [0])[0])
    except Exception:
        kpi_errors_24h = 0

    # 4) Dashboard'da tek satır "son hata" preview
    try:
        cur.execute(
            """
            SELECT COALESCE(last_error,'')
            FROM companies
            WHERE COALESCE(last_error,'') <> ''
            ORDER BY
              CASE WHEN COALESCE(last_send_at,'')<>'' THEN 0 ELSE 1 END,
              datetime(REPLACE(substr(COALESCE(last_send_at, updated_at),1,19),'T',' ')) DESC
            LIMIT 1
            """
        )
        last_error_preview = (cur.fetchone() or [""])[0].strip()
        if last_error_preview:
            last_error_preview = (last_error_preview[:160] + "…") if len(last_error_preview) > 160 else last_error_preview
        else:
            last_error_preview = ""
    except Exception:
        last_error_preview = ""

    con.close()

    return templates.TemplateResponse(
        "dashboard.html",
        {
            "request": request,
            **ctx,
            "kpi_pending": kpi_pending,
            "kpi_companies_active": kpi_companies_active,
            "kpi_errors_24h": kpi_errors_24h,
            "last_error_preview": last_error_preview,
        },
    )



@app.get("/send")
def send_center(request: Request):
    ctx = build_ctx("send", "Gönder")
    scan_err = settings_get("last_scan_err", "").strip()
    scan_msg = send_status_line() if not scan_err else ""
    stats = ready_stats_by_company()

    companies = companies_all()
    for c in companies:
        s = stats.get(c["id"], {"ready_count": 0, "ready_bytes": 0})
        c["ready_count"] = s["ready_count"]
        c["ready_human"] = format_size(s["ready_bytes"])

        if c["next_send_at"]:
            dt = parse_iso(c["next_send_at"])
            c["next_send_at_local"] = dt.astimezone(TR_TZ).strftime("%Y-%m-%d %H:%M") if dt else ""
        else:
            c["next_send_at_local"] = ""

        if c["last_send_at"]:
            dt2 = parse_iso(c["last_send_at"])
            c["last_send_at_local"] = dt2.astimezone(TR_TZ).strftime("%Y-%m-%d %H:%M") if dt2 else ""
        else:
            c["last_send_at_local"] = ""

    return templates.TemplateResponse(
        "send.html",
        {"request": request, **ctx, "companies": companies, "scan_msg": scan_msg, "scan_err": scan_err},
    )



# ===============================
# JOB TRACKING FEED (SIDEBAR + PAGE)
# ===============================
from fastapi.responses import JSONResponse
import time

# Done işleri kaç saniye görünsün (settings'ten okunacak)
def job_done_ttl_sec() -> int:
    try:
        h = int(settings_get("job_done_ttl_hours", "24") or 24)
        h = max(1, min(168, h))  # 1..168 saat
        return h * 3600
    except Exception:
        return 24 * 3600


@app.get("/job-tracking/feed")
def job_tracking_feed():
    """
    İş Takibi sayfası feed:
      - done: RECENT_JOBS (TTL ayarlı, default 24 saat)
      - sending: companies.last_send_status='SENDING'
      - planned: SADECE gelecekteki next_send_at (yazıcı kuyruğu gibi)
    Ek: geçmiş planlar otomatik DB'den temizlenir (Scheduler ile çakışmaması için 10 dk toleranslı).
    """
    now_ts = time.time()

    # --- TTL (default 24 saat) ---
    # Öncelik: job_done_ttl_hours (ayarlar sayfası) -> saniyeye çevir
    ttl_sec = 86400
    try:
        h = settings_get("job_done_ttl_hours", "").strip()
        if h:
            ttl_sec = int(h) * 3600
        else:
            ttl_sec = int(settings_get("job_done_ttl_sec", "86400") or 86400)
    except Exception:
        ttl_sec = 86400
    if ttl_sec < 60:
        ttl_sec = 60


    # --- TR format helper (dd.mm.yyyy HH:MM) ---
    def _fmt_tr(s: str) -> str:
        s = (s or "").strip()
        if not s:
            return ""
        try:
            dt = parse_iso(s)
            if dt:
                return dt.astimezone(TR_TZ).strftime("%d.%m.%Y %H:%M")
        except Exception:
            pass
        try:
            s2 = s.replace("T", " ").strip()[:16]
            dt2 = datetime.strptime(s2, "%Y-%m-%d %H:%M").replace(tzinfo=TR_TZ)
            return dt2.strftime("%d.%m.%Y %H:%M")
        except Exception:
            return s.replace("T", " ")[:16]

    # --- Geçmiş planları otomatik temizle (UTC bazında) ---
    # YETENEK KORUNDU: İşlerin arka plan görevcisiyle yarışıp kaybolmaması için 
    # saati geçen işler anında değil, 10 dakika gecikmeyle temizlenir.
    try:
        ensure_company_job_cols()
        con = db()
        cur = con.cursor()
        cur.execute(
            """
            UPDATE companies
            SET next_send_at='',
                next_send_total_files=0
            WHERE COALESCE(next_send_at,'') != ''
              AND datetime(REPLACE(substr(next_send_at,1,19),'T',' ')) < datetime('now', '-10 minutes')
              AND COALESCE(last_send_status,'') != 'SENDING'
            """
        )
        con.commit()
        con.close()
    except Exception:
        try:
            con.close()
        except Exception:
            pass

    # --- DONE (TTL) ---
    done = []
    try:
        for j in list(RECENT_JOBS):
            if now_ts - float(j.get("ts", 0)) <= ttl_sec:
                x = dict(j)
                if not x.get("when_tr"):
                    x["when_tr"] = _fmt_tr(x.get("when", ""))
                # done item'ı her zaman (Y/Y)
                if "x" not in x:
                    x["x"] = x.get("y", 0) or 0
                if "y" not in x:
                    x["y"] = x.get("x", 0) or 0
                
                x["type"] = "done"  # FRONTEND EŞLEŞMESİ İÇİN EKLENDİ
                done.append(x)
    except Exception:
        done = []

    sending = []
    planned = []

    try:
        con = db()
        cur = con.cursor()
        cur.execute(
            """
            SELECT
              id,
              name,
              next_send_at,
              COALESCE(next_send_total_files,0),
              last_send_status,
              last_send_at,
              last_send_files,
              last_error
            FROM companies
            WHERE COALESCE(next_send_at,'') != '' OR COALESCE(last_send_status,'') = 'SENDING'
            ORDER BY
              CASE WHEN COALESCE(last_send_status,'')='SENDING' THEN 0 ELSE 1 END ASC,
              CASE WHEN COALESCE(next_send_at,'')='' THEN 1 ELSE 0 END ASC,
              datetime(REPLACE(substr(next_send_at,1,19),'T',' ')) ASC,
              id DESC
            """
        )
        rows = cur.fetchall()
        con.close()

        for cid, name, next_send_at, total_files, last_st, last_at, last_files, last_err in rows:
            next_send_at = (next_send_at or "").strip()
            last_st = (last_st or "").strip()

            item = {
                "id": int(cid),
                "name": name or "-",
                "next_send_at": next_send_at,
                "next_send_at_tr": _fmt_tr(next_send_at),
                "status": last_st,
                "last_send_at_tr": _fmt_tr(last_at or ""),
                "last_send_files": int(last_files or 0),
                "last_error": (last_err or ""),
                # planned için (0/Y)
                "x": 0,
                "y": int(total_files or 0),
            }

            if last_st == "SENDING":
                item["type"] = "sending" # FRONTEND EŞLEŞMESİ İÇİN EKLENDİ
                sending.append(item)
                continue

            # planned sadece gelecekteyse
            is_future = False
            if next_send_at:
                try:
                    dtp = parse_iso(next_send_at)
                    if dtp:
                        is_future = (dtp > utcnow())
                except Exception:
                    is_future = False

            if next_send_at and is_future:
                item["type"] = "planned" # FRONTEND EŞLEŞMESİ İÇİN EKLENDİ
                planned.append(item)

    except Exception as e:
        try:
            err("JOB_TRACKING_FEED_ERR", e)
        except Exception:
            pass
        sending, planned = [], []

    # Sadece son tamamlanan iş göster (UI: tek satır)
    try:
        done = sorted(done, key=lambda x: float(x.get('ts', 0) or 0), reverse=True)[:1]
    except Exception:
        pass
    return JSONResponse({"ok": True, "sending": sending, "planned": planned, "done": done})

from fastapi.responses import JSONResponse
import json

@app.get("/jobs")
def jobs_feed():
    """
    Sidebar (legacy) feed:
      - UI tarafında bazı yerler /jobs bekliyor. 404 olmasın diye alias.
      - /job-tracking/feed çıktısını alır ve items listesi üretir.
    """
    try:
        # job_tracking_feed zaten JSONResponse dönüyor; body’yi güvenle parse edelim
        resp = job_tracking_feed()
        try:
            import json as _json
            payload = _json.loads(resp.body.decode("utf-8"))
        except Exception:
            payload = {}

        if not payload or not payload.get("ok"):
            return JSONResponse({"ok": True, "items": []})

        done = payload.get("done") or []
        sending = payload.get("sending") or []
        planned = payload.get("planned") or []

        # Sidebar items formatı: done + planned (sending’i en üste de koyabiliriz)
        # sending’i item gibi göstermek istiyorsan: type='sending' olanları başa koyuyoruz
        items = []
        items.extend(sending)
        items.extend(done)
        items.extend(planned)

        return JSONResponse({"ok": True, "items": items})
    except Exception as e:
        try:
            err("JOBS_FEED_ERR", e)
        except Exception:
            pass
        return JSONResponse({"ok": True, "items": []})

@app.get("/job-tracking")
def job_tracking_page(request: Request):
    ctx = build_ctx("job_tracking", "İş Takibi")
    return templates.TemplateResponse("job_tracking.html", {"request": request, **ctx})

@app.post("/job-tracking/cancel")
def job_tracking_cancel(company_id: int = Form(...)):
    ensure_company_job_cols()
    con = db()
    cur = con.cursor()
    cur.execute("UPDATE companies SET next_send_at='', next_send_total_files=0 WHERE id=?", (int(company_id),))
    con.commit()
    con.close()
    return RedirectResponse(url="/job-tracking?toast=İptal+Edildi", status_code=303)

@app.post("/send/scan")
def send_scan():
    """Kaynak klasörü tarar; evrakları günceller; UI durum sütununu da senkronlar."""
    scanned, matched, inserted, err = scan_source_folder()
    set_last_scan_now()

    if err:
        settings_set("last_scan_err", err)
        settings_set("last_scan_msg", "")
        return RedirectResponse(url="/send", status_code=303)

    # Scan OK
    settings_set("last_scan_err", "")
    settings_set("last_scan_msg", f"Tarandı:{scanned} | Eşleşen:{matched} | Yeni:{inserted}")

    # ✅ Durum sütunu senkronu:
    # - READY evrak varsa ve son durum boş/NO_FILES ise: READY yap
    # - READY evrak yoksa ve son durum boş/READY ise: NO_FILES yap
    try:
        stats = ready_stats_by_company()  # {cid: {ready_count, ready_bytes}}
        con = db()
        cur = con.cursor()
        cur.execute("SELECT id, COALESCE(last_send_status,'') FROM companies WHERE is_active=1")
        rows = cur.fetchall()

        for cid, last_st in rows:
            cid = int(cid)
            last_st = (last_st or "").strip()
            rc = int((stats.get(cid) or {}).get("ready_count", 0) or 0)

            if rc > 0:
                if last_st in ("", "NO_FILES"):
                    cur.execute(
                        "UPDATE companies SET last_send_status=?, last_error=? WHERE id=?",
                        ("READY", "", cid),
                    )
            else:
                if last_st in ("", "READY"):
                    cur.execute(
                        "UPDATE companies SET last_send_status=?, last_error=? WHERE id=?",
                        ("NO_FILES", "Hazır dosya yok.", cid),
                    )

        con.commit()
        con.close()
    except Exception as e:
        try:
            warn("SEND_SCAN_STATUS_SYNC_ERR", err=str(e))
        except Exception:
            pass

    return RedirectResponse(url="/send", status_code=303)




from typing import Optional
from datetime import datetime
from fastapi import Form
from fastapi.responses import RedirectResponse

@app.post("/send/plan")
def send_plan(
    company_id: int = Form(...),
    next_send_at_local: str = Form(""),
    auto_enabled: Optional[str] = Form(None),
):
    """
    Tek firma planla:
    - Tarih/saat kaydeder.
    - Checkbox (auto_enabled) geldiyse ona göre set eder.
    - Checkbox formda yoksa (None) DB’deki auto_enabled’i olduğu gibi bırakır.
    """
    next_send_at_local = (next_send_at_local or "").strip()
    next_send_at = ""

    if next_send_at_local:
        try:
            s = next_send_at_local.replace("T", " ").strip()
            dt_local = datetime.strptime(s, "%Y-%m-%d %H:%M").replace(tzinfo=TR_TZ)
            next_send_at = iso(dt_local)  # UTC ISO olarak saklanır
        except Exception:
            next_send_at = ""

    
    # Plan görünümü için (İş Takibi x/y): plan anında hazır dosya sayısını da yaz
    total_files = 0
    if next_send_at:
        try:
            rr = documents_ready_rows(int(company_id))
            total_files = sum(1 for _doc_id, _p in rr if _p and Path(str(_p)).exists())
        except Exception:
            total_files = 0
    con = db()
    cur = con.cursor()

    # checkbox formdan geldiyse set et; gelmediyse dokunma
    if auto_enabled is None:
        cur.execute(
            "UPDATE companies SET next_send_at=?, next_send_total_files=? WHERE id=?",
            (next_send_at, int(total_files or 0), int(company_id)),
        )
    else:
        auto_val = 1 if (auto_enabled == "1") else 0
        cur.execute(
            "UPDATE companies SET auto_enabled=?, next_send_at=?, next_send_total_files=? WHERE id=?",
            (auto_val, next_send_at, int(total_files or 0), int(company_id)),
        )

    con.commit()
    con.close()
    return RedirectResponse(url="/send", status_code=303)

@app.post("/companies/delete_all")
def companies_delete_all():
    con = db()
    cur = con.cursor()

    # SQLite FK varsa silmeye engel olabilir: güvenli kapat
    try:
        cur.execute("PRAGMA foreign_keys=OFF")
    except Exception:
        pass

    # Önce çocuk tablolar
    for tbl in ["documents", "companies_send_logs", "send_logs", "logs"]:
        try:
            cur.execute(f"DELETE FROM {tbl}")
        except Exception:
            pass

    # Sonra firmalar
    cur.execute("DELETE FROM companies")

    # sqlite autoincrement temizliği (varsa)
    try:
        cur.execute("DELETE FROM sqlite_sequence WHERE name IN ('companies','documents')")
    except Exception:
        pass

    con.commit()
    con.close()

    # Firmalar sayfasına dön ki gözle görünür olsun
    return RedirectResponse(url="/companies?deleted=1", status_code=303)

from fastapi import Form, Request


@app.post("/send/plan_bulk")
async def send_plan_bulk(request: Request):
    """
    Toplu planlama:
    - Formdan company_ids'yi CSV veya çoklu değer olarak alır.
    - next_send_at_local kaydeder.
    - auto_enabled checkbox geldiyse ona göre set eder.
    """
    form = await request.form()

    # company_ids: ya ["10,9,8"] gelir ya da ["10","9","8"]
    raw_list = form.getlist("company_ids")
    raw = ",".join([str(x) for x in raw_list if x is not None]).strip()

    company_ids: list[int] = []
    if raw:
        for part in raw.split(","):
            part = part.strip()
            if part.isdigit():
                company_ids.append(int(part))

    # uniq + order
    seen = set()
    company_ids = [x for x in company_ids if (x not in seen and not seen.add(x))]

    next_send_at_local = (form.get("next_send_at_local") or "").strip()
    next_send_at = ""
    if next_send_at_local:
        try:
            s = next_send_at_local.replace("T", " ").strip()
            dt_local = datetime.strptime(s, "%Y-%m-%d %H:%M").replace(tzinfo=TR_TZ)
            next_send_at = iso(dt_local)
        except Exception:
            next_send_at = ""

    auto_enabled = form.get("auto_enabled")  # "1" veya None
    auto_val = 1 if (auto_enabled == "1") else 0

    if not company_ids:
        settings_set("last_scan_err", "Seçim yok. Lütfen firmaları işaretleyin.")
        return RedirectResponse(url="/send", status_code=303)

    
    # Plan görünümü için (İş Takibi x/y): plan anında hazır dosya sayısını da yaz
    def _count_ready(cid: int) -> int:
        try:
            rr = documents_ready_rows(int(cid))
            return sum(1 for _doc_id, _p in rr if _p and Path(str(_p)).exists())
        except Exception:
            return 0

    updates = []
    for cid in company_ids:
        tf = _count_ready(int(cid)) if next_send_at else 0
        updates.append((auto_val, next_send_at, int(tf or 0), int(cid)))
    con = db()
    cur = con.cursor()
    cur.executemany(
        "UPDATE companies SET auto_enabled=?, next_send_at=?, next_send_total_files=? WHERE id=?",
        updates,
    )
    con.commit()
    con.close()

    settings_set("last_scan_msg", f"Plan kaydedildi. Firma sayısı: {len(company_ids)}")
    settings_set("last_scan_err", "")
    return RedirectResponse(url="/send?toast=Plan+Kaydedildi", status_code=303)




def _send_now_for_company(company_id: int) -> bool:
    """Endpoint dışı gönderim çekirdeği: bulk + scheduler aynı mantığı kullansın."""
    c = company_by_id(int(company_id))
    if not c:
        return False

    ok, info = wa_health()
    if not ok:
        raise RuntimeError("WhatsApp servisi hazır değil: " + str(info))

    ready_rows = documents_ready_rows(c["id"])
    existing_paths: List[str] = []
    missing_ids: List[int] = []
    for doc_id, path in ready_rows:
        if Path(path).exists():
            existing_paths.append(path)
        else:
            missing_ids.append(int(doc_id))

    if missing_ids:
        conx = db()
        cx = conx.cursor()
        now_iso = utcnow().isoformat()
        cx.executemany(
            "UPDATE documents SET status='MISSING', last_error=?, updated_at=? WHERE id=?",
            [("Dosya bulunamadı (diskte yok).", now_iso, did) for did in missing_ids],
        )
        conx.commit()
        conx.close()

    if not existing_paths:
        con0 = db()
        cur0 = con0.cursor()
        cur0.execute(
            "UPDATE companies SET last_send_status=?, last_error=?, last_send_files=? WHERE id=?",
            ("NO_FILES", "Hazır dosya yok / dosyalar kayıp.", 0, c["id"]),
        )
        con0.commit()
        con0.close()
        return False

    d = daily_get()
    lim = limits()
    planned_files = len(existing_paths)
    planned_msgs = 1 + planned_files
    planned_mb = int(sum(Path(p).stat().st_size for p in existing_paths) / (1024 * 1024))
    used_mb = int(d["bytes_sent"] / (1024 * 1024))

    if (
        d["messages_sent"] + planned_msgs > lim["daily_msg_limit"]
        or d["files_sent"] + planned_files > lim["daily_file_limit"]
        or used_mb + planned_mb > lim["daily_mb_limit"]
    ):
        con1 = db()
        cur1 = con1.cursor()
        cur1.execute(
            "UPDATE companies SET last_send_status=?, last_error=? WHERE id=?",
            ("PAUSED_LIMIT", "Günlük limit doldu", c["id"]),
        )
        con1.commit()
        con1.close()
        return False

    phone = normalize_tr_phone(c["whatsapp_phone"])
    msg_tpl = settings_get("send_message", "Vergi/SSK Tahakkuk evraklarınız ektedir.")
    msg = render_message_template(msg_tpl, c.get("name", ""), len(existing_paths))

    con2 = db()
    cur2 = con2.cursor()
    cur2.execute("UPDATE companies SET last_send_status=?, last_error=?, next_send_total_files=? WHERE id=?", ("SENDING", "", len(existing_paths), c["id"]))
    con2.commit()
    con2.close()

    ok2, info2 = wa_send_batch(
        phone,
        msg,
        existing_paths,
        company_id=c["id"],
        company_name=c.get("name", ""),
    )
    if not ok2:
        con3 = db()
        cur3 = con3.cursor()
        cur3.execute("UPDATE companies SET last_send_status=?, last_error=? WHERE id=?", ("ERROR", str(info2), c["id"]))
        con3.commit()
        con3.close()
        raise RuntimeError(str(info2))

    moved, aerr = archive_ready_docs(c)
    status = "SENT" if not aerr else "SENT_ARCHIVE_ERROR"

    total_bytes = 0
    for p in existing_paths:
        try:
            total_bytes += Path(p).stat().st_size
        except Exception:
            pass
    daily_add(messages=1 + len(existing_paths), files=len(existing_paths), bytes_=total_bytes)

    # Ek sayacı + status text
    _mark_company_sent_with_extra(c["id"], existing_paths, status, aerr, extra_label="Ek")

    try:
        dbg("SEND_DONE", company_id=c["id"], name=c.get("name",""), files=len(existing_paths), moved=moved, aerr=aerr or "")
    except Exception:
        pass

    return True


@app.post("/send/instant_bulk")
async def send_instant_bulk(request: Request):
    """Toplu hemen gönder: seçili firmalar için instant send tetikler (dosya yoksa atlar).
    Not: Aynı anda tek gönderim çalışır. Meşgulse seçilenler kuyruğa alınır.
    """
    form = await request.form()

    # company_ids ham formdan: ['10,9,8'] veya ['10','9',...]
    raw_list = form.getlist("company_ids")
    raw = ",".join([str(x) for x in raw_list if x is not None]).strip()

    company_ids: list[int] = []
    if raw:
        for part in raw.split(","):
            part = part.strip()
            if part.isdigit():
                company_ids.append(int(part))

    # uniq + order
    seen = set()
    company_ids = [x for x in company_ids if (x not in seen and not seen.add(x))]

    if not company_ids:
        settings_set("last_scan_err", "Seçim yok. Lütfen firmaları işaretleyin.")
        return RedirectResponse(url="/send", status_code=303)

    ok, info = wa_health()
    if not ok:
        settings_set("last_scan_err", "WhatsApp servisi hazır değil: " + info)
        return RedirectResponse(url="/send", status_code=303)

    # Meşgulse: hepsini kuyruğa al (hızlı dönüş)
    if SEND_MUTEX.locked():
        base = utcnow() + timedelta(seconds=30)
        queued = 0
        for k, cid in enumerate(company_ids):
            try:
                c = company_by_id(int(cid))
                if not c:
                    continue
                ready_rows = documents_ready_rows(int(cid))
                existing = [p for _did, p in ready_rows if p and Path(p).exists()]
                when_iso = iso(base + timedelta(seconds=20 * k))
                _queue_company_send(int(cid), when_iso, len(existing), note="Toplu gönderim kuyruğa alındı.")
                queued += 1
            except Exception:
                continue

        settings_set("last_scan_err", "")
        settings_set("last_scan_msg", f"Toplu gönderim kuyruğa alındı. Firma sayısı: {queued}")
        return RedirectResponse(url="/send?toast=Kuyruğa+Alındı", status_code=303)

    acquired = SEND_MUTEX.acquire(blocking=False)
    if not acquired:
        settings_set("last_scan_err", "Sistem meşgul. Tekrar deneyin.")
        return RedirectResponse(url="/send", status_code=303)

    sent = 0
    skipped = 0
    try:
        for cid in company_ids:
            c = company_by_id(int(cid))
            if not c:
                skipped += 1
                continue
            try:
                _send_now_for_company(int(cid))
                sent += 1
                # Çok hızlı üst üste olmasın
                await asyncio.sleep(1.0)
            except Exception as e:
                skipped += 1
                log_event(int(cid), c.get("name", ""), "ERROR", f"Toplu hemen gönder hata: {e}", 0)

        settings_set("last_scan_msg", f"Toplu gönderim tamamlandı. Gönderilen: {sent} | Atlanan: {skipped}")
        settings_set("last_scan_err", "")
        return RedirectResponse(url="/send?toast=Toplu+Gönderim", status_code=303)
    finally:
        try:
            SEND_MUTEX.release()
        except Exception:
            pass
@app.post("/send/instant")
def send_instant(company_id: int = Form(...)):
    c = company_by_id(int(company_id))
    if not c:
        settings_set("last_scan_err", "Firma bulunamadı.")
        return RedirectResponse(url="/send", status_code=303)

    ok, info = wa_health()
    if not ok:
        settings_set("last_scan_err", "WhatsApp servisi hazır değil: " + info)
        return RedirectResponse(url="/send", status_code=303)

    ready_rows = documents_ready_rows(c["id"])
    existing_paths: List[str] = []
    missing_ids: List[int] = []
    for doc_id, path in ready_rows:
        if Path(path).exists():
            existing_paths.append(path)
        else:
            missing_ids.append(int(doc_id))

    if missing_ids:
        conx = db()
        cx = conx.cursor()
        now_iso = utcnow().isoformat()
        cx.executemany(
            "UPDATE documents SET status='MISSING', last_error=?, updated_at=? WHERE id=?",
            [("Dosya bulunamadı (diskte yok).", now_iso, did) for did in missing_ids],
        )
        conx.commit()
        conx.close()

    if not existing_paths:
        con0 = db()
        cur0 = con0.cursor()
        cur0.execute(
            "UPDATE companies SET last_send_status=?, last_error=?, last_send_files=? WHERE id=?",
            ("NO_FILES", "Hazır dosya yok / dosyalar kayıp.", 0, c["id"]),
        )
        con0.commit()
        con0.close()
        settings_set("last_scan_err", "")
        settings_set("last_scan_msg", f"{c['name']} için hazır dosya yok. Gönderim yapılmadı.")
        return RedirectResponse(url="/send", status_code=303)

    d = daily_get()
    lim = limits()
    planned_files = len(existing_paths)
    planned_msgs = 1 + planned_files
    planned_mb = int(sum(Path(p).stat().st_size for p in existing_paths) / (1024 * 1024))
    used_mb = int(d["bytes_sent"] / (1024 * 1024))

    if (
        d["messages_sent"] + planned_msgs > lim["daily_msg_limit"]
        or d["files_sent"] + planned_files > lim["daily_file_limit"]
        or used_mb + planned_mb > lim["daily_mb_limit"]
    ):
        con1 = db()
        cur1 = con1.cursor()
        cur1.execute(
            "UPDATE companies SET last_send_status=?, last_error=? WHERE id=?",
            ("PAUSED_LIMIT", "Günlük limit doldu", c["id"]),
        )
        con1.commit()
        con1.close()
        settings_set("last_scan_err", "Günlük limit doldu, gönderim durduruldu.")
        return RedirectResponse(url="/send", status_code=303)

    phone = normalize_tr_phone(c["whatsapp_phone"])
    msg_tpl = settings_get("send_message", "Vergi/SSK Tahakkuk evraklarınız ektedir.")
    msg = render_message_template(msg_tpl, c.get("name", ""), len(existing_paths))

    # Aynı anda tek gönderim: meşgulse kuyruğa al
    acquired = SEND_MUTEX.acquire(blocking=False)
    if not acquired:
        # Kuyruğa al (yaklaşık 30 sn sonra)
        when_iso = iso(utcnow() + timedelta(seconds=30))
        _queue_company_send(int(company_id), when_iso, len(existing_paths), note="Kuyruğa alındı (gönderim devam ediyor).")
        settings_set("last_scan_err", "")
        settings_set("last_scan_msg", f"{c['name']} için gönderim kuyruğa alındı.")
        return RedirectResponse(url="/send?toast=Kuyruğa+Alındı", status_code=303)

    try:
        con2 = db()
        cur2 = con2.cursor()
        cur2.execute("UPDATE companies SET last_send_status=?, last_error=?, next_send_total_files=? WHERE id=?", ("SENDING", "", len(existing_paths), c["id"]))
        con2.commit()
        con2.close()

        ok2, info2 = wa_send_batch(phone, msg, existing_paths, company_id=int(company_id), company_name=c.get("name",""))
        if not ok2:
            con3 = db()
            cur3 = con3.cursor()
            cur3.execute("UPDATE companies SET last_send_status=?, last_error=? WHERE id=?", ("ERROR", info2, c["id"]))
            con3.commit()
            con3.close()
            settings_set("last_scan_err", f"{c['name']} gönderim hatası: {info2}")
            return RedirectResponse(url="/send", status_code=303)

        moved, aerr = archive_ready_docs(c)
        status = "SENT" if not aerr else "SENT_ARCHIVE_ERROR"

        total_bytes = 0
        for p in existing_paths:
            try:
                total_bytes += Path(p).stat().st_size
            except Exception:
                pass
        daily_add(messages=1 + len(existing_paths), files=len(existing_paths), bytes_=total_bytes)

        # ✅ başarılı: şirket kaydı + Ek sayacı + plan temizliği
        _mark_company_sent_with_extra(int(c["id"]), existing_paths, status, aerr, company_name=c.get("name",""))

        push_recent_job_done(
            company_id=int(c["id"]),
            company_name=c.get("name",""),
            files=len(existing_paths),
        )

    finally:
        try:
            SEND_MUTEX.release()
        except Exception:
            pass
    settings_set("last_scan_err", "")
    settings_set("last_scan_msg", f"{c['name']} ✅ gönderildi. Dosya:{len(existing_paths)} Arşiv:{moved}")
    return RedirectResponse(url="/send", status_code=303)

@app.get("/settings")
def settings_page(request: Request):
    ctx = build_ctx("settings", "Ayarlar")

    settings = {
        "source_dir": settings_get("source_dir", ""),
        "archive_dir": settings_get("archive_dir", ""),
        "send_message": settings_get("send_message", "Vergi/SSK Tahakkuk evraklarınız ektedir."),
        "per_file_min_s": settings_get("per_file_min_s", "8"),
        "per_file_max_s": settings_get("per_file_max_s", "15"),
        "daily_msg_limit": settings_get("daily_msg_limit", "1100"),
        "daily_file_limit": settings_get("daily_file_limit", "1100"),
        "daily_mb_limit": settings_get("daily_mb_limit", "1500"),
    }

    # settings.html: value="{{ job_done_ttl_hours or 24 }}" bekliyor
    try:
        ttl_sec = int(settings_get("job_done_ttl_sec", "86400") or 86400)
    except Exception:
        ttl_sec = 86400
    if ttl_sec < 3600:
        ttl_sec = 3600
    if ttl_sec > 168 * 3600:
        ttl_sec = 168 * 3600
    job_done_ttl_hours = int(ttl_sec // 3600)

    return templates.TemplateResponse(
        "settings.html",
        {
            "request": request,
            **ctx,
            "settings": settings,
            "job_done_ttl_hours": job_done_ttl_hours,
        },
    )

@app.post("/settings/save")
def settings_save(
    request: Request,
    source_dir: str = Form(""),
    archive_dir: str = Form(""),
    send_message: str = Form(""),
    per_file_min_s: str = Form("8"),
    per_file_max_s: str = Form("15"),
    daily_msg_limit: str = Form("1100"),
    daily_file_limit: str = Form("1100"),
    daily_mb_limit: str = Form("1500"),
    job_done_ttl_hours: str = Form("24"),
):
    source_dir = (source_dir or "").strip()
    archive_dir = (archive_dir or "").strip()
    if source_dir and (not archive_dir):
        archive_dir = str(Path(source_dir) / "arsiv")

    settings_set("source_dir", source_dir)
    settings_set("archive_dir", archive_dir)
    settings_set("send_message", (send_message or "").strip())
    settings_set("per_file_min_s", str((per_file_min_s or "8").strip()))
    settings_set("per_file_max_s", str((per_file_max_s or "15").strip()))
    settings_set("daily_msg_limit", str((daily_msg_limit or "1100").strip()))
    settings_set("daily_file_limit", str((daily_file_limit or "1100").strip()))
    settings_set("daily_mb_limit", str((daily_mb_limit or "1500").strip()))

    # İş Takibi TTL saat -> saniye
    try:
        h = int(str(job_done_ttl_hours or "24").strip())
    except Exception:
        h = 24
    if h < 1:
        h = 1
    if h > 168:
        h = 168
    settings_set("job_done_ttl_sec", str(h * 3600))

    settings_set("last_scan_err", "")
    settings_set("last_scan_msg", "Ayarlar kaydedildi.")
    return RedirectResponse(url="/settings?toast=Kaydedildi", status_code=303)


@app.get("/companies")
def companies_page(request: Request):
    ctx = build_ctx("companies", "Firmalar")
    return templates.TemplateResponse("companies.html", {"request": request, **ctx, "companies": companies_all()})


@app.post("/settings/job-ttl")
def settings_job_ttl(job_done_ttl_hours: int = Form(...)):
    """
    İş Takibi 'Tamamlananlar' görünme süresi (saat).
    Default: 24 saat
    """
    try:
        h = int(job_done_ttl_hours)
    except Exception:
        h = 24

    # güvenli aralık: 1 saat - 168 saat (7 gün)
    if h < 1:
        h = 1
    if h > 168:
        h = 168

    settings_set("job_done_ttl_sec", str(h * 3600))
    settings_set("last_scan_msg", f"İş Takibi görünme süresi güncellendi: {h} saat")
    settings_set("last_scan_err", "")
    return RedirectResponse(url="/settings?toast=Kaydedildi", status_code=303)


@app.post("/companies/add")
def companies_add(
    request: Request,
    name: str = Form(...),
    match_key: str = Form(...),
    whatsapp_phone: str = Form(...),
    is_active: str = Form("1"),
):
    name = (name or "").strip()
    match_key = (match_key or "").strip()
    phone = normalize_tr_phone(whatsapp_phone or "")
    if not name or not match_key or not phone:
        return RedirectResponse(url="/companies", status_code=303)
    con = db()
    cur = con.cursor()
    cur.execute(
        "INSERT INTO companies(name, match_key, whatsapp_phone, is_active, created_at) VALUES(?,?,?,?,?)",
        (name, match_key, phone, int(is_active or "1"), iso(utcnow())),
    )
    con.commit()
    con.close()
    return RedirectResponse(url="/companies", status_code=303)


@app.post("/companies/update")
def companies_update(
    company_id: int = Form(...),
    name: str = Form(...),
    match_key: str = Form(...),
    whatsapp_phone: str = Form(...),
    is_active: str = Form("1"),
):
    con = db()
    cur = con.cursor()
    cur.execute(
        """
      UPDATE companies
      SET name=?, match_key=?, whatsapp_phone=?, is_active=?
      WHERE id=?
    """,
        ((name or "").strip(), (match_key or "").strip(), normalize_tr_phone(whatsapp_phone or ""), int(is_active or "1"), int(company_id)),
    )
    con.commit()
    con.close()
    return RedirectResponse(url="/companies", status_code=303)


@app.get("/companies/export")
def companies_export():
    wb = Workbook()
    ws = wb.active
    ws.title = "Firmalar"
    ws.append(["firma_adi", "eslestirme_anahtari", "telefon", "aktif", "auto_enabled", "next_send_at_local"])
    for c in companies_all():
        next_local = ""
        if c["next_send_at"]:
            dt = parse_iso(c["next_send_at"])
            if dt:
                next_local = dt.astimezone(TR_TZ).strftime("%Y-%m-%d %H:%M")
        ws.append([c["name"], c["match_key"], c["whatsapp_phone"], c["is_active"], c["auto_enabled"], next_local])
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=firmalar.xlsx"},
    )


@app.post("/companies/import")
async def companies_import(file: UploadFile = File(...)):
    content = await file.read()
    wb = load_workbook(filename=io.BytesIO(content))
    ws = wb.active
    header = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(header)}

    def cell(row, key, default=""):
        i = idx.get(key, None)
        if i is None:
            return default
        v = row[i].value
        return str(v).strip() if v is not None else default

    con = db()
    cur = con.cursor()

    for row in ws.iter_rows(min_row=2):
        name = cell(row, "firma_adi")
        mk = cell(row, "eslestirme_anahtari")
        phone = normalize_tr_phone(cell(row, "telefon"))
        aktif = cell(row, "aktif", "1")
        auto_enabled = cell(row, "auto_enabled", "0")
        next_local = cell(row, "next_send_at_local", "")

        if not name or not mk or not phone:
            continue

        next_send_at = ""
        if next_local:
            try:
                s = next_local.replace('T',' ').strip()
                dt_local = datetime.strptime(s, "%Y-%m-%d %H:%M").replace(tzinfo=TR_TZ)
                next_send_at = iso(dt_local)
            except Exception:
                next_send_at = ""

        cur.execute("SELECT id FROM companies WHERE match_key=?", (mk,))
        existing = cur.fetchone()
        if existing:
            cur.execute(
                """
              UPDATE companies
              SET name=?, whatsapp_phone=?, is_active=?, auto_enabled=?, next_send_at=?
              WHERE match_key=?
            """,
                (name, phone, int(aktif or "1"), int(auto_enabled or "0"), next_send_at, mk),
            )
        else:
            cur.execute(
                """
              INSERT INTO companies(name, match_key, whatsapp_phone, is_active, created_at, auto_enabled, next_send_at)
              VALUES(?,?,?,?,?,?,?)
            """,
                (name, mk, phone, int(aktif or "1"), iso(utcnow()), int(auto_enabled or "0"), next_send_at),
            )

    con.commit()
    con.close()
    return RedirectResponse(url="/companies", status_code=303)


@app.get("/logs")
def logs_page(request: Request):
    ctx = build_ctx("logs", "Loglar")
    scan_err = settings_get("last_scan_err", "").strip()
    scan_msg = send_status_line() if not scan_err else ""

    con = db()
    cur = con.cursor()
    cur.execute(
    """
    SELECT id, name, last_send_status, last_send_at, last_error, last_send_files
    FROM companies
    WHERE COALESCE(last_send_status,'') != '' OR COALESCE(last_error,'') != ''
    ORDER BY
      CASE WHEN last_send_at IS NULL OR TRIM(last_send_at) = '' THEN 1 ELSE 0 END ASC,
      datetime(REPLACE(last_send_at,'T',' ')) DESC,
      id DESC
    """
    )
    rows = cur.fetchall()
    con.close()

    recent = []
    for cid, name, st, last_at, err, last_files in rows:
        dt = parse_iso(last_at) if last_at else None
        last_tr = dt.astimezone(TR_TZ).strftime("%Y-%m-%d %H:%M") if dt else (last_at or "-")
        st = st or ""
        st_tr = {
            "SENT": "Gönderildi",
            "DONE": "Gönderildi",
            "SENDING": "Gönderiliyor",
            "PAUSED_LIMIT": "Limit",
            "NO_FILES": "Dosya yok",
            "ERROR": "Hata",
        }.get(st, st or "-")
        recent.append({"id": cid, "name": name, "status_tr": st_tr, "last_at": last_tr, "err": err or "", "files": int(last_files or 0)})

    return templates.TemplateResponse(
        "logs.html",
        {"request": request, **ctx, "scan_err": scan_err, "scan_msg": scan_msg, "last_scan": last_scan_text(), "recent": recent},
    )
from fastapi.responses import JSONResponse

@app.post("/logs/clear")
def logs_clear():
    con = db()
    cur = con.cursor()

    cur.execute(
        """
        UPDATE companies
        SET
          last_send_status = '',
          last_send_at = '',
          last_error = '',
          last_send_files = 0
        WHERE COALESCE(last_send_status,'') != '' OR COALESCE(last_error,'') != ''
        """
    )

    con.commit()
    con.close()
    return JSONResponse({"ok": True})

from fastapi import Query

@app.get("/archive")
def archive_page(
    request: Request,
    company: str = Query(default=""),
    from_: str = Query(default="", alias="from"),
    to: str = Query(default=""),
):
    ctx = build_ctx("archive", "Arşiv")

    where = ["d.status='ARCHIVED'"]
    params = []

    if company.strip():
        where.append("LOWER(COALESCE(c.name,'')) LIKE ?")
        params.append(f"%{company.strip().lower()}%")

    # ISO string ise ilk 10 karakter YYYY-MM-DD
    if from_.strip():
        where.append("substr(COALESCE(d.updated_at,''),1,10) >= ?")
        params.append(from_.strip())

    if to.strip():
        where.append("substr(COALESCE(d.updated_at,''),1,10) <= ?")
        params.append(to.strip())

    sql = f"""
      SELECT d.id, d.filename, d.size_bytes, d.updated_at, c.name
      FROM documents d
      LEFT JOIN companies c ON c.id = d.company_id
      WHERE {" AND ".join(where)}
      ORDER BY
        CASE WHEN d.updated_at IS NULL OR TRIM(d.updated_at) = '' THEN 1 ELSE 0 END ASC,
        datetime(REPLACE(substr(d.updated_at,1,19),'T',' ')) DESC,
        d.id DESC
      LIMIT 500
    """

    con = db()
    cur = con.cursor()
    cur.execute(sql, params)
    rows = cur.fetchall()
    con.close()

    items = []
    for doc_id, filename, size_bytes, updated_at, company_name in rows:
        dt = parse_iso(updated_at) if updated_at else None
        items.append(
            {
                "id": doc_id,
                "company": company_name or "-",
                "filename": filename,
                "size": format_size(int(size_bytes or 0)),
                "archived_at": (dt.astimezone(TR_TZ).strftime("%Y-%m-%d %H:%M") if dt else (updated_at or "-")),
            }
        )

    return templates.TemplateResponse(
        "archive.html",
        {
            "request": request,
            **ctx,
            "items": items,
            "q_company": company.strip(),
            "q_from": from_.strip(),
            "q_to": to.strip(),
        },
    )

from fastapi.responses import StreamingResponse
import io

@app.get("/archive/export")
def archive_export(
    company: str = Query(default=""),
    from_: str = Query(default="", alias="from"),
    to: str = Query(default=""),
    format: str = Query(default="xlsx"),
):
    # sadece xlsx destekle
    fmt = (format or "xlsx").lower().strip()
    if fmt != "xlsx":
        fmt = "xlsx"

    where = ["d.status='ARCHIVED'"]
    params = []

    if company.strip():
        where.append("LOWER(COALESCE(c.name,'')) LIKE ?")
        params.append(f"%{company.strip().lower()}%")

    if from_.strip():
        where.append("substr(COALESCE(d.updated_at,''),1,10) >= ?")
        params.append(from_.strip())

    if to.strip():
        where.append("substr(COALESCE(d.updated_at,''),1,10) <= ?")
        params.append(to.strip())

    sql = f"""
      SELECT d.id, COALESCE(c.name,'-') as company, d.filename, d.size_bytes, d.updated_at
      FROM documents d
      LEFT JOIN companies c ON c.id = d.company_id
      WHERE {" AND ".join(where)}
      ORDER BY
        CASE WHEN d.updated_at IS NULL OR TRIM(d.updated_at) = '' THEN 1 ELSE 0 END ASC,
        datetime(REPLACE(substr(d.updated_at,1,19),'T',' ')) DESC,
        d.id DESC
      LIMIT 500
    """

    con = db()
    cur = con.cursor()
    cur.execute(sql, params)
    rows = cur.fetchall()
    con.close()

    # openpyxl ile xlsx üret
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Arsiv"

    ws.append(["No", "Firma", "Dosya", "Boyut (byte)", "Arşiv Tarihi"])
    for doc_id, comp, filename, size_bytes, updated_at in rows:
        ws.append([doc_id, comp, filename, int(size_bytes or 0), updated_at or ""])

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="arsiv.xlsx"'},
    )

from fastapi.responses import JSONResponse

from pathlib import Path
import os

@app.post("/archive/clear")
def archive_clear(
    company: str = Query(default=""),
    from_: str = Query(default="", alias="from"),
    to: str = Query(default=""),
):
    where = ["d.status='ARCHIVED'"]
    params = []

    if company.strip():
        where.append("LOWER(COALESCE(c.name,'')) LIKE ?")
        params.append(f"%{company.strip().lower()}%")

    if from_.strip():
        where.append("substr(COALESCE(d.updated_at,''),1,10) >= ?")
        params.append(from_.strip())

    if to.strip():
        where.append("substr(COALESCE(d.updated_at,''),1,10) <= ?")
        params.append(to.strip())

    con = db()
    cur = con.cursor()

    # id + source_path al
    cur.execute(
        f"""
        SELECT d.id, d.source_path
        FROM documents d
        LEFT JOIN companies c ON c.id = d.company_id
        WHERE {" AND ".join(where)}
        """,
        params,
    )
    rows = cur.fetchall()

    deleted = 0
    ids = []

    for doc_id, p in rows:
        ids.append(doc_id)
        # diskten sil (path varsa ve mevcutsa)
        try:
            if p:
                fp = Path(p)
                if fp.exists() and fp.is_file():
                    fp.unlink()
        except:
            pass  # disk silme hatası DB silmeyi engellemesin

    if ids:
        q = ",".join(["?"] * len(ids))
        cur.execute(f"DELETE FROM documents WHERE id IN ({q})", ids)
        deleted = len(ids)

    con.commit()
    con.close()
    return JSONResponse({"ok": True, "deleted": deleted})

@app.get("/api/pick-folder")
def api_pick_folder():
    try:
        import tkinter as tk
        from tkinter import filedialog

        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        path = filedialog.askdirectory(title="Klasör Seç")
        root.destroy()
        return {"ok": True, "path": path or ""}
    except Exception as e:
        return {"ok": False, "error": str(e), "path": ""}

# ===============================
# DEBUG READY DOCS + DEBUG LOGGING (FINAL)
# Bu bloğu main.py EN SONUNA ekle (tek parça)
# ===============================

# --- debug endpoint için gerekli importlar ---
from pathlib import Path
from datetime import datetime
from typing import Any, Optional

from fastapi import Query, Request, HTTPException
from fastapi.responses import JSONResponse
from fastapi.exceptions import RequestValidationError

import json
import traceback


@app.get("/debug/ready_docs")
def debug_ready_docs(company_id: int = Query(...)):
    """
    Seçili firmanın READY evrak yollarını ve disk varlığını gösterir.
    """
    rows = documents_ready_rows(company_id)  # (doc_id, path)
    out = []
    for doc_id, path in rows:
        p = str(path or "")
        out.append(
            {
                "id": int(doc_id),
                "path": p,
                "exists": Path(p).exists() if p else False,
            }
        )
    return JSONResponse({"company_id": int(company_id), "count": len(out), "items": out})


# ===============================
# DEBUG / WARNING / ERROR LOGGING
# ===============================

_BASE_DIR = Path(__file__).resolve().parent
_LOG_DIR = _BASE_DIR / "logs"
_LOG_DIR.mkdir(parents=True, exist_ok=True)
_LOG_FILE = _LOG_DIR / "app_debug.log"


def _now_tr() -> str:
    try:
        return datetime.now(TR_TZ).strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _safe_json(x: Any) -> str:
    try:
        return json.dumps(x, ensure_ascii=False, default=str)
    except Exception:
        return str(x)


def _write_line(line: str) -> None:
    # console
    try:
        print(line)
    except Exception:
        pass

    # file append (utf-8)
    try:
        with open(_LOG_FILE, "a", encoding="utf-8") as f:
            f.write(line + "\n")
    except Exception:
        pass


def dbg(msg: str, **kv: Any) -> None:
    line = f"[{_now_tr()}] DEBUG {msg}"
    if kv:
        line += " | " + _safe_json(kv)
    _write_line(line)


def warn(msg: str, **kv: Any) -> None:
    line = f"[{_now_tr()}] WARN  {msg}"
    if kv:
        line += " | " + _safe_json(kv)
    _write_line(line)


def err(msg: str, exc: Optional[BaseException] = None, **kv: Any) -> None:
    line = f"[{_now_tr()}] ERROR {msg}"
    if kv:
        line += " | " + _safe_json(kv)
    _write_line(line)
    if exc is not None:
        tb = "".join(traceback.format_exception(type(exc), exc, exc.__traceback__))
        _write_line(tb.strip())


def activity(name: str, **kv: Any) -> None:
    """Faaliyet başlangıcı / durum logu için kısa helper."""
    dbg(f"ACT:{name}", **kv)


def ui_msg(text: str) -> None:
    """UI'da görünen last_scan_msg'e de yaz (varsa) + logla"""
    try:
        settings_set("last_scan_msg", text)
    except Exception:
        pass
    dbg("UI_MSG", text=text)


def ui_err(text: str) -> None:
    """UI'da görünen last_scan_err'e de yaz (varsa) + logla"""
    try:
        settings_set("last_scan_err", text)
    except Exception:
        pass
    err("UI_ERR", None, text=text)


def db_log(company_id: int, company_name: str, status: str, message: str, files: int = 0) -> None:
    """log_event varsa DB'ye de düş"""
    try:
        log_event(int(company_id), company_name or "", status, message, int(files))
    except Exception:
        dbg("DB_LOG_FALLBACK", company_id=company_id, status=status, message=message, files=files)


# ===============================
# Global exception handlers
# - 422 validation hataları 422 olarak kalsın
# - HTTPException'lar bozulmasın
# - Diğerleri 500 + log
# ===============================

from starlette.exceptions import HTTPException as StarletteHTTPException
from fastapi import HTTPException as FastAPIHTTPException
from fastapi.exceptions import RequestValidationError
from fastapi import Request
from fastapi.responses import JSONResponse

@app.exception_handler(RequestValidationError)
async def _validation_exception_handler(request: Request, exc: RequestValidationError):
    warn("VALIDATION_ERROR", path=str(request.url.path), method=request.method, detail=str(exc)[:800])
    return JSONResponse(status_code=422, content={"detail": exc.errors()})

@app.exception_handler(Exception)
async def _global_exception_handler(request: Request, exc: Exception):
    # ✅ 405/404 gibi starlette HTTPException'ları bozma
    if isinstance(exc, (StarletteHTTPException, FastAPIHTTPException)):
        raise exc

    err("UNHANDLED_EXCEPTION", exc, path=str(request.url.path), method=request.method)
    return JSONResponse(status_code=500, content={"ok": False, "error": "Sunucu hatası. Loglara yazıldı."})